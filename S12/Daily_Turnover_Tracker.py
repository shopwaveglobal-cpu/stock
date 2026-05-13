#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Daily Turnover Tracker - 매일 거래대금 5000억+ 종목 추적
- 매일 1회 실행하여 오늘의 거래대금 순위 수집
- 5000억 이상 종목만 엑셀에 축적
- (날짜, 티커) 중복 자동 제거
"""

import os
import sys
import argparse
import logging
import time
from datetime import datetime, date
from typing import List, Dict, Tuple, Optional

import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# 거래일 체크 유틸리티 import
try:
    from trading_day_utils import is_trading_day, get_trading_day_info
    TRADING_DAY_CHECK_AVAILABLE = True
except ImportError:
    TRADING_DAY_CHECK_AVAILABLE = False
    logger_import = logging.getLogger(__name__)
    logger_import.warning("trading_day_utils 모듈을 찾을 수 없습니다. 거래일 체크가 비활성화됩니다.")

# ==================== 설정 ====================
APPKEY_DEFAULT = "IweTdkYa8JWDUOa8NohVSVeOiJ1THDGd_2x050A8XcU"
SECRET_DEFAULT = "eazu-jPNJpAsIVkaUTh3_88gUvXrCMJCwGF2AYRtBJs"

API_BASE_URL = "https://api.kiwoom.com"
API_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
API_RANK_ENDPOINT = "/api/dostk/rkinfo"
API_RANK_ID = "ka10032"
API_CHART_ENDPOINT = "/api/dostk/chart"  # 양봉 필터용
API_CHART_ID = "ka10081"  # 양봉 필터용

EXCEL_PATH = "output/turnover_universe.xlsx"
SHEET_NAME = "universe"

THRESHOLD_EOK = 5000.0  # 5000억
MARKETS = ["001", "101"]  # KRX, KOSDAQ

# ETF/ETN 제외 키워드
EXCLUDE_KEYWORDS = [
    "KODEX", "TIGER", "KBSTAR", "KOSEF", "ARIRANG", "HANARO", "SOL", 
    "TREX", "ACE", "인버스", "레버리지", "선물", "ETF", "ETN", "지수"
]

# ==================== 로깅 ====================
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger(__name__)

# ==================== API 함수 ====================
def get_access_token(appkey: str, secret: str) -> str:
    """키움 API 토큰 획득"""
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    body = {
        "grant_type": "client_credentials",
        "appkey": appkey.strip('"'),
        "secretkey": secret.strip('"')
    }
    
    response = requests.post(API_TOKEN_URL, headers=headers, json=body, timeout=20)
    response.raise_for_status()
    
    data = response.json()
    token = data.get("token") or data.get("access_token")
    
    if not token:
        raise RuntimeError("토큰 획득 실패")
    
    logger.info("✓ API 토큰 획득 완료")
    return token


def fetch_rank_data(token: str, market: str, max_retry: int = 5) -> dict:
    """거래대금 순위 조회 (재시도 포함)"""
    headers = {
        "authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "api-id": API_RANK_ID,
        "cont-yn": "N",
        "next-key": ""
    }
    
    body = {
        "qry_tp": "2",          # 전일 거래대금 상위
        "rank_strt": "0",
        "rank_end": "100",      # 상위 100개
        "stex_tp": "3",         # 통합
        "mang_stk_incls": "1",
        "mrkt_tp": market,
        "bas_dd": datetime.now().strftime("%Y%m%d")  # 오늘 날짜 (실제론 최신만 반환)
    }
    
    url = API_BASE_URL + API_RANK_ENDPOINT
    
    for attempt in range(max_retry):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=20)
            
            # Rate limit 처리
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After", 1)
                sleep_time = float(retry_after) if str(retry_after).isdigit() else (0.5 * (2 ** attempt))
                logger.warning(f"Rate limit - {sleep_time:.1f}초 대기 중...")
                time.sleep(sleep_time)
                continue
            
            # 서버 오류 재시도
            if 500 <= response.status_code < 600:
                logger.warning(f"서버 오류 {response.status_code} - 재시도 {attempt + 1}/{max_retry}")
                time.sleep(0.5 * (2 ** attempt))
                continue
            
            response.raise_for_status()
            return response.json()
            
        except requests.RequestException as e:
            if attempt == max_retry - 1:
                raise
            logger.warning(f"요청 실패 - 재시도 {attempt + 1}/{max_retry}: {e}")
            time.sleep(0.5 * (2 ** attempt))
    
    raise RuntimeError("최대 재시도 횟수 초과")


def fetch_today_candle(token: str, ticker: str, max_retry: int = 3) -> Optional[Dict[str, float]]:
    """
    당일 캔들 데이터 조회 (양봉 필터용)

    Args:
        token: API 인증 토큰
        ticker: 종목 티커 (6자리)
        max_retry: 최대 재시도 횟수

    Returns:
        {"open": 시가, "close": 종가} 또는 None (실패 시)
    """
    # 통합 종목코드
    integrated_ticker = f"{ticker}_AL"

    headers = {
        "authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "api-id": API_CHART_ID,
        "cont-yn": "N",
        "next-key": ""
    }

    body = {
        "stk_cd": integrated_ticker,
        "base_dt": datetime.now().strftime("%Y%m%d"),
        "upd_stkpc_tp": "1",  # 수정주가
        "stex_tp": "3"  # 통합
    }

    url = API_BASE_URL + API_CHART_ENDPOINT

    for attempt in range(max_retry):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=20)
            response.raise_for_status()
            result = response.json()

            # 데이터 추출
            records = None
            for value in result.values():
                if isinstance(value, list) and len(value) > 0:
                    records = value
                    break

            if not records:
                logger.warning(f"  차트 데이터 없음: {ticker}")
                return None

            # 첫 번째 레코드 (최신 데이터)
            first_record = records[0]

            # 시가/종가 파싱
            open_price = float(first_record.get("open_pric", 0))
            close_price = float(first_record.get("cur_prc", 0))

            if open_price > 0 and close_price > 0:
                return {"open": open_price, "close": close_price}
            else:
                logger.warning(f"  유효하지 않은 가격 데이터: {ticker}")
                return None

        except Exception as e:
            if attempt == max_retry - 1:
                logger.warning(f"  차트 조회 실패: {ticker} - {e}")
                return None
            time.sleep(0.3 * (2 ** attempt))

    return None


def is_bullish_candle(token: str, ticker: str) -> bool:
    """
    양봉 여부 확인 (시가 < 종가)

    Args:
        token: API 인증 토큰
        ticker: 종목 티커

    Returns:
        True: 양봉, False: 음봉 또는 조회 실패
    """
    candle = fetch_today_candle(token, ticker)

    if candle is None:
        return False

    # 양봉: 시가 < 종가
    return candle["open"] < candle["close"]


# ==================== 데이터 처리 ====================
def normalize_ticker(ticker: str) -> str:
    """티커 정규화 (6자리, 알파벳 포함 가능)"""
    if not ticker:
        return ""
    
    # _AL 같은 suffix 제거
    ticker = str(ticker).split("_")[0].strip()
    
    # 알파벳이 포함된 경우 (예: 0008Z0) 그대로 반환
    if any(c.isalpha() for c in ticker):
        return ticker.zfill(6)[:6]
    
    # 숫자만 있는 경우 숫자만 추출하여 6자리로 패딩
    digits = "".join(c for c in ticker if c.isdigit())
    return digits.zfill(6)[:6] if digits else ""


def is_excluded(name: str) -> bool:
    """ETF/ETN 등 제외 대상 여부"""
    if not name:
        return True
    
    name_upper = str(name).upper()
    return any(keyword.upper() in name_upper for keyword in EXCLUDE_KEYWORDS)


def parse_rank_response(response: dict) -> List[Dict]:
    """API 응답에서 종목 정보 추출"""
    # 응답에서 리스트 데이터 찾기
    data_list = None
    for value in response.values():
        if isinstance(value, list) and len(value) > 0:
            data_list = value
            break
    
    if not data_list:
        return []
    
    results = []
    
    for item in data_list:
        # 티커 추출
        ticker = None
        for field in ["stk_cd", "stck_shrn_iscd", "ISU_SRT_CD", "isu_cd"]:
            if field in item and item[field]:
                ticker = normalize_ticker(item[field])
                break
        
        # 종목명 추출
        name = None
        for field in ["stk_kor_isnm", "stk_kor_nm", "ISU_ABBRV", "stk_nm", "isu_nm"]:
            if field in item and item[field]:
                name = str(item[field]).strip()
                break
        
        # 거래대금 추출
        turnover = None
        for field in ["tvol_tamt", "TOT_TR_PRC", "tot_tr_prc", "trdval", "trd_prc", "trde_prica", "trdval_amt"]:
            if field in item and item[field]:
                try:
                    turnover = float(str(item[field]).replace(",", ""))
                except (ValueError, TypeError):
                    continue
                break
        
        # 유효한 데이터만 추가
        if ticker and name and turnover:
            results.append({
                "ticker": ticker,
                "name": name,
                "turnover_100man": turnover  # 100만원 단위
            })
    
    return results


def filter_stocks(data: List[Dict], threshold_eok: float) -> pd.DataFrame:
    """거래대금 필터링 및 정렬"""
    if not data:
        return pd.DataFrame(columns=["ticker", "name", "turnover_eok"])
    
    df = pd.DataFrame(data)
    
    # 티커 기준 중복 제거 (거래대금 합산)
    df = df.groupby(["ticker", "name"], as_index=False)["turnover_100man"].sum()
    
    # 단위 변환: 100만원 → 억원 (×0.01)
    df["turnover_eok"] = df["turnover_100man"] * 0.01
    
    # ETF/ETN 제외
    df = df[~df["name"].apply(is_excluded)].copy()
    
    # 거래대금 임계값 필터
    df = df[df["turnover_eok"] >= threshold_eok].copy()
    
    # 거래대금 내림차순 정렬
    df = df.sort_values("turnover_eok", ascending=False).reset_index(drop=True)
    
    return df[["ticker", "name", "turnover_eok"]]


# ==================== 엑셀 처리 ====================
def ensure_excel_exists(path: str):
    """엑셀 파일이 없으면 생성"""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"])
        wb.save(path)
        logger.info(f"✓ 새 엑셀 파일 생성: {path}")


def read_existing_data(path: str) -> pd.DataFrame:
    """기존 엑셀 데이터 읽기 (시트 이름 자동 감지)"""
    ensure_excel_exists(path)

    try:
        # 먼저 'universe' 시트로 시도
        df = pd.read_excel(path, sheet_name=SHEET_NAME, dtype={"티커": str})
    except ValueError as e:
        # 'universe' 시트가 없으면 첫 번째 시트 읽기
        if "Worksheet" in str(e) and "not found" in str(e):
            logger.warning(f"'{SHEET_NAME}' 시트를 찾을 수 없습니다. 첫 번째 시트를 읽습니다.")
            try:
                df = pd.read_excel(path, sheet_name=0, dtype={"티커": str})
                logger.info(f"✓ 첫 번째 시트 읽기 성공 ({len(df)}개 종목)")
            except Exception as e2:
                logger.error(f"✗ 첫 번째 시트 읽기 실패: {e2}")
                return pd.DataFrame(columns=["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"])
        else:
            logger.error(f"엑셀 읽기 실패: {e}")
            return pd.DataFrame(columns=["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"])
    except Exception as e:
        logger.error(f"엑셀 읽기 실패: {e}")
        return pd.DataFrame(columns=["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"])
    
    if df.empty:
        return df
    
    # Unnamed 컬럼들과 업데이트 관련 컬럼 제거
    cols_to_drop = [col for col in df.columns if "Unnamed" in str(col) or "업데이트" in str(col) or "최종" in str(col)]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    
    # 구버전 호환성: "날짜" → "첫주도주", "최근주도주"
    if "날짜" in df.columns and "첫주도주" not in df.columns:
        df["첫주도주"] = df["날짜"]
        df["최근주도주"] = df["날짜"]
        df = df.drop(columns=["날짜"])
    
    # 누적횟수 열이 없으면 추가
    if "누적횟수" not in df.columns:
        df["누적횟수"] = 1
    
    # 데이터 정규화
    df["티커"] = df["티커"].apply(normalize_ticker)
    df["첫주도주"] = pd.to_datetime(df["첫주도주"], errors="coerce").dt.date
    df["최근주도주"] = pd.to_datetime(df["최근주도주"], errors="coerce").dt.date
    df["누적횟수"] = df["누적횟수"].fillna(1).astype(int)
    df = df.dropna(subset=["첫주도주"])
    
    # ⭐ 컬럼 순서 정리 (신버전 표준 순서로)
    expected_cols = ["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"]
    df = df[expected_cols]
    
    return df


def get_last_update_date(path: str) -> Optional[date]:
    """H1 셀에서 마지막 업데이트 날짜 읽기 (시트 이름 자동 감지)"""
    try:
        wb = load_workbook(path)

        # 시트 이름 자동 감지
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            # 첫 번째 시트 사용
            ws = wb.worksheets[0]

        h1_value = ws["H1"].value
        
        if h1_value and isinstance(h1_value, str):
            # "최종 업데이트: 2025-10-12" 형태에서 날짜 추출
            if ":" in h1_value:
                # "최종 업데이트: " 제거 후 날짜 파싱
                parts = h1_value.split(":", 1)  # 첫 번째 :로만 분리
                if len(parts) > 1:
                    date_str = parts[1].strip()  # "2025-10-12"
                    return pd.to_datetime(date_str).date()
        
        return None
    except Exception as e:
        return None


def save_to_excel(path: str, df: pd.DataFrame, update_date: str = None):
    """데이터프레임을 openpyxl로 직접 저장 (pandas ExcelWriter 사용 안 함)"""
    from openpyxl import Workbook

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # 헤더 작성
    headers = ["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)

    # 데이터 작성
    for idx, row in df.iterrows():
        ws.cell(idx + 2, 1, row["첫주도주"])  # date 객체 그대로 저장
        ws.cell(idx + 2, 2, row["최근주도주"])  # date 객체 그대로 저장
        ws.cell(idx + 2, 3, row["티커"])
        ws.cell(idx + 2, 4, row["종목명"])
        ws.cell(idx + 2, 5, row["거래대금(억)"])
        ws.cell(idx + 2, 6, row["누적횟수"])

    # 서식 적용 (워크시트 객체 전달)
    apply_formatting_to_ws(ws, update_date)

    # 파일 저장
    wb.save(path)


def apply_formatting_to_ws(ws, update_date: str = None):
    """워크시트에 서식 적용 (열 너비, 테두리, 정렬, 숫자 포맷)"""
    try:
        from openpyxl.styles import Alignment, Font

        # 열 너비 자동 조절 (컬럼 내용에 따라)
        from openpyxl.utils import get_column_letter

        # 각 컬럼의 최대 길이를 계산하여 너비 자동 설정
        for col_idx in range(1, 7):  # A~F열
            col_letter = get_column_letter(col_idx)
            max_length = 0

            # 헤더 길이 확인
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))

            # 데이터 길이 확인
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length

            # 너비 설정 (최소값과 여유 공간 추가)
            adjusted_width = min(max(max_length + 2, 8), 50)  # 최소 8, 최대 50
            ws.column_dimensions[col_letter].width = adjusted_width

        # 테두리
        thin_border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # ⭐ 헤더 행 서식 (테두리 + 중앙정렬)
        for col_idx in range(1, 7):  # A~F열 헤더
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ⭐ 데이터 행 서식 (값이 있는 셀만 테두리 적용)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)

                # 값이 있는 셀만 처리
                if cell.value not in (None, ""):
                    cell.border = thin_border

                    # A,B열(날짜): 날짜 포맷 + 중앙정렬 (⭐ 시간 제거, 날짜만 표시)
                    if col_idx in [1, 2]:
                        cell.number_format = 'yyyy-mm-dd'
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # C열(티커): 텍스트 포맷 + 중앙정렬 (⭐ 앞의 0 보존)
                    elif col_idx == 3:
                        cell.number_format = '@'
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # D열(종목명): 중앙정렬
                    elif col_idx == 4:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # E열(거래대금): 천단위 콤마 + 오른쪽 정렬
                    elif col_idx == 5:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

                    # F열(누적횟수): 정수 + 오른쪽 정렬
                    elif col_idx == 6:
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

        # ⭐ H1 셀에 업데이트 날짜 기록 (테두리 없음, 왼쪽 정렬)
        if update_date:
            ws["H1"] = f"최종 업데이트: {update_date}"
            ws["H1"].font = Font(bold=True, size=10)
            ws["H1"].alignment = Alignment(horizontal="left", vertical="center")
            ws["H1"].border = None  # 테두리 없음

    except Exception as e:
        logger.warning(f"서식 적용 실패: {e}")


def append_to_excel(path: str, new_rows: List[Tuple[date, str, str, float]]):
    """새 데이터를 엑셀에 추가 (종목별 1행 유지 + 누적횟수 카운트)"""
    if not new_rows:
        logger.info("저장할 데이터가 없습니다.")
        return
    
    # 오늘 날짜 (스크립트 실행 날짜)
    today = date.today()
    
    # H1 셀에서 마지막 업데이트 날짜 읽기
    last_update = get_last_update_date(path)
    
    # 기존 데이터 읽기
    df_old = read_existing_data(path)
    
    # 새 데이터 생성 (API에서 받은 거래일)
    df_new = pd.DataFrame(new_rows, columns=["거래일", "티커", "종목명", "거래대금(억)"])
    df_new["티커"] = df_new["티커"].apply(normalize_ticker)
    trading_date = pd.to_datetime(df_new["거래일"]).dt.date.iloc[0]  # API가 준 거래일
    
    # 신규 종목용: 첫주도주 = 최근주도주 = 거래일
    df_new["첫주도주"] = trading_date
    df_new["최근주도주"] = trading_date
    df_new["누적횟수"] = 1
    df_new = df_new.drop(columns=["거래일"])  # 임시 컬럼 제거
    
    # 기존 종목과 신규 종목 분리
    if not df_old.empty:
        existing_tickers = set(df_old["티커"].values)
        
        # 기존에 있는 종목: 업데이트
        mask_existing = df_new["티커"].isin(existing_tickers)
        df_existing_update = df_new[mask_existing].copy()
        
        # 기존에 없는 종목: 그대로 추가
        df_really_new = df_new[~mask_existing].copy()
        
        # 기존 데이터 업데이트
        same_day_count = 0  # 같은 날 재실행 카운트
        
        for _, row in df_existing_update.iterrows():
            ticker = row["티커"]
            mask = df_old["티커"] == ticker
            
            # ⭐ 핵심: H1 셀의 업데이트날짜 기준으로 판단 (거래일이 아님!)
            # 오늘 처음 실행이면 (H1의 날짜가 오늘이 아니면) 누적횟수 +1
            if last_update != today:
                df_old.loc[mask, "누적횟수"] += 1
            else:
                # 오늘 이미 실행했었음 (같은 날 재실행)
                same_day_count += 1
            
            # 첫주도주: 유지 (변경 안 함)
            # 최근주도주: 항상 최신 거래일로 갱신
            df_old.loc[mask, "최근주도주"] = row["최근주도주"]
            df_old.loc[mask, "거래대금(억)"] = row["거래대금(억)"]
        
        # 기존 데이터 + 신규 종목만 합치기
        df_all = pd.concat([df_old, df_really_new], ignore_index=True)
        
        new_count = len(df_really_new)
        updated_count = len(df_existing_update)
    else:
        # 기존 데이터가 없으면 모두 신규
        df_all = df_new
        new_count = len(df_new)
        updated_count = 0
        same_day_count = 0
    
    # 누적횟수 내림차순 → 최근주도주 내림차순 정렬
    df_all = df_all.sort_values(["누적횟수", "최근주도주"], ascending=[False, False]).reset_index(drop=True)
    
    # H1 셀에 기록할 업데이트 날짜
    update_date = today.strftime("%Y-%m-%d")
    
    # 저장 (H1 셀 포함)
    save_to_excel(path, df_all, update_date)
    
    # 결과 로그
    if same_day_count > 0 and same_day_count == updated_count:
        logger.info(f"✓ 데이터 저장 완료: {updated_count}개 종목 갱신 (오늘 이미 실행함, 누적횟수 유지)")
    elif new_count > 0 and updated_count > 0:
        actual_increase = updated_count - same_day_count
        if actual_increase > 0:
            logger.info(f"✓ 데이터 저장 완료: 신규 {new_count}개, 누적횟수 증가 {actual_increase}개" + 
                       (f", 재실행 {same_day_count}개" if same_day_count > 0 else ""))
        else:
            logger.info(f"✓ 데이터 저장 완료: 신규 {new_count}개")
    elif new_count > 0:
        logger.info(f"✓ 데이터 저장 완료: 신규 {new_count}개 종목 추가")
    elif updated_count > 0:
        actual_increase = updated_count - same_day_count
        if actual_increase > 0:
            logger.info(f"✓ 데이터 저장 완료: {actual_increase}개 종목의 누적횟수 증가" +
                       (f" (재실행 {same_day_count}개 제외)" if same_day_count > 0 else ""))
        else:
            logger.info(f"✓ 데이터 저장 완료: {updated_count}개 종목 갱신 (오늘 이미 실행함)")
    else:
        logger.info("✓ 데이터 저장 완료: 변경 없음")


# ==================== 메인 로직 ====================
def collect_today_data(token: str, threshold_eok: float, excel_path: str, filter_bullish: bool = False) -> int:
    """오늘의 거래대금 데이터 수집 및 저장"""
    today = date.today()
    logger.info(f"{'='*60}")
    logger.info(f"날짜: {today}")
    logger.info(f"임계값: {threshold_eok:,.0f}억 이상")
    if filter_bullish:
        logger.info(f"필터: 양봉 종목만 (시가 < 종가)")
    logger.info(f"{'='*60}")
    
    # 모든 마켓 데이터 수집
    all_data = []
    
    for market in MARKETS:
        market_name = "KRX" if market == "001" else "KOSDAQ"
        logger.info(f"[{market_name}] 거래대금 순위 조회 중...")
        
        try:
            response = fetch_rank_data(token, market)
            parsed = parse_rank_response(response)
            all_data.extend(parsed)
            logger.info(f"  ✓ {len(parsed)}개 종목 수신")
        except Exception as e:
            logger.error(f"  ✗ 조회 실패: {e}")
    
    if not all_data:
        logger.warning("조회된 데이터가 없습니다.")
        return 0
    
    # 필터링
    df_filtered = filter_stocks(all_data, threshold_eok)

    if df_filtered.empty:
        logger.warning(f"{threshold_eok:,.0f}억 이상 종목이 없습니다.")
        return 0

    # 양봉 필터 적용 (선택적)
    if filter_bullish:
        logger.info(f"\n양봉 필터링 진행 중... (총 {len(df_filtered)}개 종목)")
        bullish_stocks = []

        for idx, row in df_filtered.iterrows():
            ticker = row['ticker']
            name = row['name']

            if is_bullish_candle(token, ticker):
                bullish_stocks.append(idx)
                logger.info(f"  ✓ {name} ({ticker}): 양봉")
            else:
                logger.info(f"  ✗ {name} ({ticker}): 음봉 또는 조회 실패")

        # 양봉 종목만 남기기
        df_filtered = df_filtered.loc[bullish_stocks]

        if df_filtered.empty:
            logger.warning("양봉 종목이 없습니다.")
            return 0

    # 결과 출력
    logger.info(f"\n{'='*60}")
    logger.info(f"거래대금 {threshold_eok:,.0f}억 이상: {len(df_filtered)}개 종목")
    if filter_bullish:
        logger.info(f"(양봉 필터 적용됨)")
    logger.info(f"{'='*60}")
    
    logger.info("\n상위 10개:")
    for idx, row in df_filtered.head(10).iterrows():
        logger.info(f"  {idx+1:2d}. {row['ticker']} {row['name']:20s} {row['turnover_eok']:>10,.0f}억")
    
    # 엑셀에 저장
    rows_to_save = [
        (today, row["ticker"], row["name"], row["turnover_eok"])
        for _, row in df_filtered.iterrows()
    ]
    
    append_to_excel(excel_path, rows_to_save)
    
    return len(rows_to_save)


# ==================== 엔트리 포인트 ====================
def main():
    parser = argparse.ArgumentParser(
        description="매일 거래대금 5000억+ 종목 추적 스크립트"
    )
    parser.add_argument(
        "--threshold", 
        type=float, 
        default=THRESHOLD_EOK,
        help=f"거래대금 임계값 (억원, 기본값: {THRESHOLD_EOK:,.0f})"
    )
    parser.add_argument(
        "--out", 
        default=EXCEL_PATH,
        help=f"출력 엑셀 파일명 (기본값: {EXCEL_PATH})"
    )
    parser.add_argument(
        "--appkey",
        default=None,
        help="키움 API APPKEY (환경변수 KIWOOM_APPKEY 또는 기본값 사용)"
    )
    parser.add_argument(
        "--secret",
        default=None,
        help="키움 API SECRET (환경변수 KIWOOM_SECRET 또는 기본값 사용)"
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="상세 로그 출력"
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="거래일 체크 무시하고 강제 실행"
    )
    parser.add_argument(
        "--no-filter-bullish",
        action="store_true",
        help="양봉 필터 비활성화 (기본값: 양봉 필터 활성화)"
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # 거래일 체크 (강제 실행 옵션이 없는 경우에만)
    if TRADING_DAY_CHECK_AVAILABLE and not args.force:
        trading_info = get_trading_day_info()
        if not trading_info['is_trading_day']:
            logger.info("=" * 60)
            logger.info(f"📅 비거래일입니다 ({trading_info['reason']})")
            logger.info("거래일이 아닌 날에는 데이터 수집을 건너뜁니다.")
            logger.info("강제 실행하려면 --force 옵션을 사용하세요.")
            logger.info("=" * 60)
            return
    elif args.force:
        logger.info("🔧 강제 실행 모드: 거래일 체크를 무시합니다.")
    
    # API 키 설정
    appkey = args.appkey or os.getenv("KIWOOM_APPKEY") or APPKEY_DEFAULT
    secret = args.secret or os.getenv("KIWOOM_SECRET") or SECRET_DEFAULT
    
    excel_path = os.path.abspath(args.out)
    
    try:
        # 토큰 획득 (환경변수로 전달된 경우 재발급 생략)
        token = os.getenv("KIWOOM_TOKEN") or get_access_token(appkey, secret)

        # 데이터 수집 및 저장 (기본값: 양봉 필터 활성화)
        filter_bullish = not args.no_filter_bullish
        count = collect_today_data(token, args.threshold, excel_path, filter_bullish=filter_bullish)

        logger.info(f"\n{'='*60}")
        logger.info(f"완료: {count}개 종목 저장됨")
        logger.info(f"파일: {excel_path}")
        logger.info(f"{'='*60}")

    except Exception as e:
        logger.exception(f"오류 발생: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

