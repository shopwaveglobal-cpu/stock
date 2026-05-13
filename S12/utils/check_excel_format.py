#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 파일 구조 확인 스크립트"""

from openpyxl import load_workbook

EXCEL_PATH = "output/turnover_universe.xlsx"

wb = load_workbook(EXCEL_PATH)
ws = wb['universe']

print("=" * 60)
print("Excel 파일 구조 확인")
print("=" * 60)
print(f"시트명: {ws.title}")
print(f"최대 컬럼: {ws.max_column}")
print(f"최대 행: {ws.max_row}")

print("\n헤더 행 (1~10열):")
for i in range(1, min(ws.max_column + 1, 10)):
    val = ws.cell(1, i).value
    print(f"  {i}열: [{val}]")

print("\nG1-J1 셀 값:")
for col in ['G', 'H', 'I', 'J']:
    val = ws[f'{col}1'].value
    print(f"  {col}1: [{val}]")

print("\n첫 번째 데이터 행 (1~8열):")
for i in range(1, 9):
    val = ws.cell(2, i).value
    print(f"  {i}열: [{val}]")
