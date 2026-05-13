#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
11/6~11/7 손실 데이터 복원 스크립트
로그에서 거래대금 데이터를 추출하여 Excel에 반영
"""

import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

EXCEL_PATH = "output/turnover_universe.xlsx"

# 로그에서 추출한 데이터
data_20251106 = [
    ("000660", "SK하이닉스", 54671),
    ("005930", "삼성전자", 48358),
    ("034020", "티센크루프엘리베이터", 16232),
    ("035420", "NAVER", 12862),
    ("486990", "메타", 12813),
    ("098460", "고영", 12561),
    ("278470", "펄어비스", 8711),
    ("196170", "알테오젠", 8184),
    ("086520", "에코프로", 5434),
    ("006400", "삼성SDI", 5189),
]

data_20251107 = [
    ("000660", "SK하이닉스", 47354),
    ("005930", "삼성전자", 39675),
    ("034020", "티센크루프엘리베이터", 13224),
    ("486990", "메타", 12705),
    ("035420", "NAVER", 7576),
    ("035720", "카카오", 6116),
    ("042660", "한화오션", 5530),
    ("006400", "삼성SDI", 5236),
]

print("=" * 60)
print("손실 데이터 복원 스크립트")
print("=" * 60)

# 기존 데이터 읽기
try:
    # 첫 번째 시트 읽기 (시트 이름 무관)
    df_old = pd.read_excel(EXCEL_PATH, sheet_name=0, dtype={"티커": str})
    print(f"\n[OK] 기존 데이터 읽기 성공: {len(df_old)}개 종목")
except Exception as e:
    print(f"[ERROR] 파일 읽기 실패: {e}")
    exit(1)

# Unnamed 컬럼 제거
cols_to_drop = [col for col in df_old.columns if "Unnamed" in str(col) or "업데이트" in str(col) or "최종" in str(col)]
if cols_to_drop:
    df_old = df_old.drop(columns=cols_to_drop)

# 날짜 컬럼을 datetime으로 변환 (일관성 유지)
df_old["첫주도주"] = pd.to_datetime(df_old["첫주도주"])
df_old["최근주도주"] = pd.to_datetime(df_old["최근주도주"])

# 11/6 데이터 처리
print("\n[11/6 데이터 복원]")
for ticker, name, turnover_eok in data_20251106:
    ticker = ticker.zfill(6)

    if ticker in df_old["티커"].values:
        # 기존 종목: 최근주도주 갱신, 누적횟수 +1
        mask = df_old["티커"] == ticker
        old_date = df_old.loc[mask, "최근주도주"].iloc[0]

        # 11/6보다 이전 데이터만 갱신
        if old_date < pd.to_datetime("2025-11-06"):
            df_old.loc[mask, "최근주도주"] = pd.to_datetime("2025-11-06")
            df_old.loc[mask, "거래대금(억)"] = turnover_eok
            df_old.loc[mask, "누적횟수"] += 1
            print(f"  갱신: {name} ({ticker}) - 최근주도주: {old_date.date()} → 2025-11-06")
    else:
        # 신규 종목 추가
        new_row = {
            "첫주도주": pd.to_datetime("2025-11-06"),
            "최근주도주": pd.to_datetime("2025-11-06"),
            "티커": ticker,
            "종목명": name,
            "거래대금(억)": turnover_eok,
            "누적횟수": 1
        }
        df_old = pd.concat([df_old, pd.DataFrame([new_row])], ignore_index=True)
        print(f"  신규: {name} ({ticker})")

# 11/7 데이터 처리
print("\n[11/7 데이터 복원]")
for ticker, name, turnover_eok in data_20251107:
    ticker = ticker.zfill(6)

    if ticker in df_old["티커"].values:
        # 기존 종목: 최근주도주 갱신, 누적횟수 +1
        mask = df_old["티커"] == ticker
        old_date = df_old.loc[mask, "최근주도주"].iloc[0]

        # 11/7보다 이전 데이터만 갱신
        if old_date < pd.to_datetime("2025-11-07"):
            df_old.loc[mask, "최근주도주"] = pd.to_datetime("2025-11-07")
            df_old.loc[mask, "거래대금(억)"] = turnover_eok
            df_old.loc[mask, "누적횟수"] += 1
            print(f"  갱신: {name} ({ticker}) - 최근주도주: {old_date.date()} → 2025-11-07")
    else:
        # 신규 종목 추가
        new_row = {
            "첫주도주": pd.to_datetime("2025-11-07"),
            "최근주도주": pd.to_datetime("2025-11-07"),
            "티커": ticker,
            "종목명": name,
            "거래대금(억)": turnover_eok,
            "누적횟수": 1
        }
        df_old = pd.concat([df_old, pd.DataFrame([new_row])], ignore_index=True)
        print(f"  신규: {name} ({ticker})")

# 누적횟수 내림차순 → 최근주도주 내림차순 정렬
df_old = df_old.sort_values(["누적횟수", "최근주도주"], ascending=[False, False]).reset_index(drop=True)

print(f"\n[OK] 총 {len(df_old)}개 종목")

# openpyxl로 저장
print("\n데이터 저장 중...")

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "universe"

# 헤더
headers = ["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"]
for col_idx, header in enumerate(headers, start=1):
    ws.cell(1, col_idx, header)

# 데이터
for idx, row in df_old.iterrows():
    ws.cell(idx + 2, 1, row["첫주도주"].date() if pd.notna(row["첫주도주"]) else None)
    ws.cell(idx + 2, 2, row["최근주도주"].date() if pd.notna(row["최근주도주"]) else None)
    ws.cell(idx + 2, 3, row["티커"])
    ws.cell(idx + 2, 4, row["종목명"])
    ws.cell(idx + 2, 5, row["거래대금(억)"])
    ws.cell(idx + 2, 6, row["누적횟수"])

# 서식 적용
thin_border = Border(
    top=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# 열 너비
for col_idx in range(1, 7):
    col_letter = get_column_letter(col_idx)
    max_length = 0

    header_cell = ws.cell(row=1, column=col_idx)
    if header_cell.value:
        max_length = len(str(header_cell.value))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

    adjusted_width = min(max(max_length + 2, 8), 50)
    ws.column_dimensions[col_letter].width = adjusted_width

# 헤더 서식
for col_idx in range(1, 7):
    cell = ws.cell(row=1, column=col_idx)
    if cell.value:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 데이터 서식
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, 7):
        cell = ws.cell(row=row_idx, column=col_idx)

        if cell.value not in (None, ""):
            cell.border = thin_border

            if col_idx in [1, 2]:  # 날짜
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:  # 티커
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:  # 종목명
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 5:  # 거래대금
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 6:  # 누적횟수
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

# H1 셀에 업데이트 날짜
today = date.today().strftime("%Y-%m-%d")
ws["H1"] = f"최종 업데이트: {today}"
ws["H1"].font = Font(bold=True, size=10)
ws["H1"].alignment = Alignment(horizontal="left", vertical="center")
ws["H1"].border = None

wb.save(EXCEL_PATH)
print(f"[OK] 파일 저장 완료: {EXCEL_PATH}")

print("\n" + "=" * 60)
print("복원 완료!")
print("=" * 60)
