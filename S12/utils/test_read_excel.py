#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 파일 읽기 테스트"""

import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH = "output/turnover_universe.xlsx"

print("=" * 60)
print("Excel 파일 읽기 테스트")
print("=" * 60)

# pandas로 읽기
df = pd.read_excel(EXCEL_PATH, sheet_name='universe', dtype={'티커': str})
print(f"\n[Pandas] 데이터 읽기 성공: {len(df)}행")
print(f"컬럼: {list(df.columns[:6])}")
print(f"\n첫 번째 행 데이터:")
for col in df.columns[:6]:
    print(f"  {col}: {df[col].iloc[0]}")

# openpyxl로 읽기
wb = load_workbook(EXCEL_PATH)
ws = wb['universe']
print(f"\n[Openpyxl] max_column: {ws.max_column}, max_row: {ws.max_row}")
print(f"H1 셀: {ws['H1'].value}")
print("\n정상적으로 데이터를 읽을 수 있습니다!")
