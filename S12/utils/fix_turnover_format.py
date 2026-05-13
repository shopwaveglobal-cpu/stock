#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기존 turnover_universe.xlsx 파일의 양식을 수정하는 유틸리티
- Unnamed 컬럼 제거
- 시트 이름 정규화 (Sheet1 → universe)
- 날짜 포맷 수정 (시간 제거)
- 테두리 적용 (값이 있는 셀만)
- H1 셀 업데이트 날짜 (테두리 없음, 왼쪽 정렬)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import date

EXCEL_PATH = "output/turnover_universe.xlsx"
SHEET_NAME = "universe"


def fix_turnover_format():
    """turnover_universe.xlsx 파일 양식 수정"""

    print("=" * 60)
    print("turnover_universe.xlsx 양식 수정 시작")
    print("=" * 60)

    # 1. pandas로 데이터 읽기 (기본 시트)
    try:
        df = pd.read_excel(EXCEL_PATH)
        print(f"\n[OK] 파일 읽기 성공: {len(df)}개 행")
        print(f"현재 컬럼: {df.columns.tolist()}")
    except Exception as e:
        print(f"[ERROR] 파일 읽기 실패: {e}")
        return

    # 2. Unnamed 컬럼 제거 및 정리
    cols_to_drop = [col for col in df.columns if "Unnamed" in str(col) or "업데이트" in str(col) or "최종" in str(col)]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
        print(f"[OK] 불필요한 컬럼 제거: {cols_to_drop}")

    # 3. 컬럼 이름 정규화 (혹시 모를 깨진 인코딩 수정)
    expected_cols = ["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"]
    df.columns = expected_cols[:len(df.columns)]
    print(f"[OK] 컬럼 정규화 완료: {df.columns.tolist()}")

    # 4. 날짜 컬럼을 datetime으로 변환 (나중에 Excel에서 포맷 적용)
    df["첫주도주"] = pd.to_datetime(df["첫주도주"], errors="coerce")
    df["최근주도주"] = pd.to_datetime(df["최근주도주"], errors="coerce")
    print("[OK] 날짜 포맷 변환 완료")

    # 5. 티커를 문자열로 변환 (앞의 0 보존)
    df["티커"] = df["티커"].astype(str).str.zfill(6)

    # 6. openpyxl로 직접 저장 (pandas ExcelWriter 사용하지 않음)
    print("\nopenpyxl로 직접 저장 중...")
    save_with_openpyxl(df)

    print("\n" + "=" * 60)
    print("[COMPLETE] 양식 수정 완료!")
    print("=" * 60)


def save_with_openpyxl(df):
    """openpyxl로 직접 데이터 저장 및 서식 적용"""
    from openpyxl import Workbook

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # 헤더 작성
    headers = ["첫주도주", "최근주도주", "티커", "종목명", "거래대금(억)", "누적횟수"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)

    # 데이터 작성
    for row_idx, row in df.iterrows():
        ws.cell(row_idx + 2, 1, row["첫주도주"])  # datetime 객체 그대로 저장
        ws.cell(row_idx + 2, 2, row["최근주도주"])  # datetime 객체 그대로 저장
        ws.cell(row_idx + 2, 3, row["티커"])
        ws.cell(row_idx + 2, 4, row["종목명"])
        ws.cell(row_idx + 2, 5, row["거래대금(억)"])
        ws.cell(row_idx + 2, 6, row["누적횟수"])

    print(f"[OK] 데이터 저장 완료 ({len(df)}행, 시트명: {SHEET_NAME})")

    # 서식 적용
    print("[OK] 서식 적용 중...")
    apply_formatting_to_ws(ws)

    # 파일 저장
    wb.save(EXCEL_PATH)
    print("[OK] 파일 저장 완료")


def apply_formatting_to_ws(ws):
    """워크시트에 서식 적용"""

    # 열 너비 자동 조절
    for col_idx in range(1, 7):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        # 헤더 길이
        header_cell = ws.cell(row=1, column=col_idx)
        if header_cell.value:
            max_length = len(str(header_cell.value))

        # 데이터 길이
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

        adjusted_width = min(max(max_length + 2, 8), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 테두리
    thin_border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # 헤더 행 서식 (테두리 + 중앙정렬)
    for col_idx in range(1, 7):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 행 서식
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)

            if cell.value not in (None, ""):
                cell.border = thin_border

                # A,B열(날짜): yyyy-mm-dd 포맷 + 중앙정렬
                if col_idx in [1, 2]:
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # C열(티커): 텍스트 포맷 + 중앙정렬
                elif col_idx == 3:
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # D열(종목명): 중앙정렬
                elif col_idx == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # E열(거래대금): 천단위 콤마 + 오른쪽 정렬
                elif col_idx == 5:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                # F열(누적횟수): 정수 + 오른쪽 정렬
                elif col_idx == 6:
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                # 빈 셀은 테두리 제거
                cell.border = None

    # H1 셀에 업데이트 날짜 (테두리 없음, 왼쪽 정렬)
    today = date.today().strftime("%Y-%m-%d")
    ws["H1"] = f"최종 업데이트: {today}"
    ws["H1"].font = Font(bold=True, size=10)
    ws["H1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["H1"].border = None

    print("[OK] 서식 적용 완료")
    print("  - 테두리: 값이 있는 셀만 적용")
    print("  - 날짜: yyyy-mm-dd 포맷 (시간 제거)")
    print("  - H1 셀: 테두리 없음, 왼쪽 정렬")


if __name__ == "__main__":
    fix_turnover_format()
