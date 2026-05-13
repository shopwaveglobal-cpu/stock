import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

print("=== 가격 데이터 천 단위 콤마 적용 ===")

# 1. trading_signals.xlsx 읽기
df_summary = pd.read_excel('output/trading_signals.xlsx', sheet_name='Summary')
df_history = pd.read_excel('output/trading_signals.xlsx', sheet_name='History')

print(f"Summary: {len(df_summary)}개 종목")
print(f"History: {len(df_history)}개 종목")

# 2. 가격 관련 컬럼들에 천 단위 콤마 적용
price_columns = [
    '종가', '저가', '고가', '20일선', '-20%엔벨로프',
    '1차매수선', '1차매수가', '2차매수선', '2차매수가', 
    '3차매수선', '3차매수가', '평균매수가',
    '1차매도선(+3%)', '2차매도선(+5%)', '3차매도선(+7%)', '최고도달선'
]

# Summary에서 가격 컬럼들을 문자열로 변환 (콤마 포함)
for col in price_columns:
    if col in df_summary.columns:
        df_summary[col] = df_summary[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) and x != 0 else "")

# History에서도 동일하게 적용
for col in price_columns:
    if col in df_history.columns:
        df_history[col] = df_history[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) and x != 0 else "")

print("가격 데이터 천 단위 콤마 적용 완료")

# 3. Excel 파일로 저장 (포맷팅 포함)
with pd.ExcelWriter('output/trading_signals.xlsx', engine='openpyxl') as writer:
    df_summary.to_excel(writer, index=False, sheet_name='Summary')
    df_history.to_excel(writer, index=False, sheet_name='History')

# 4. Excel 포맷팅 적용
wb = load_workbook('output/trading_signals.xlsx')

# Summary 시트 포맷팅
ws_summary = wb['Summary']

# 헤더 찾기
headers = [cell.value for cell in ws_summary[1]]

# 테두리 스타일
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 데이터 행 포맷팅
for row_idx in range(2, ws_summary.max_row + 1):
    for col_idx in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        header = headers[col_idx - 1]
        
        # 테두리 적용
        cell.border = thin_border
        
        # 가격 컬럼은 오른쪽 정렬
        if header in price_columns:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        # 이격도 컬럼은 중앙 정렬
        elif '이격도' in str(header):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # 날짜 컬럼은 중앙 정렬
        elif '일' in str(header):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # 기타는 왼쪽 정렬
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# 헤더 포맷팅
for col_idx in range(1, ws_summary.max_column + 1):
    header_cell = ws_summary.cell(row=1, column=col_idx)
    header_cell.border = thin_border
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

# 열 너비 자동 조정
for col_idx in range(1, ws_summary.max_column + 1):
    column_letter = get_column_letter(col_idx)
    header = headers[col_idx - 1]
    
    if header == "티커":
        ws_summary.column_dimensions[column_letter].width = 10
    elif header == "종목명":
        ws_summary.column_dimensions[column_letter].width = 15
    elif header == "상태메시지":
        ws_summary.column_dimensions[column_letter].width = 30
    elif header in price_columns:
        ws_summary.column_dimensions[column_letter].width = 15
    else:
        ws_summary.column_dimensions[column_letter].width = 12

# History 시트도 동일하게 포맷팅
if len(df_history) > 0:
    ws_history = wb['History']
    headers_history = [cell.value for cell in ws_history[1]]
    
    for row_idx in range(2, ws_history.max_row + 1):
        for col_idx in range(1, ws_history.max_column + 1):
            cell = ws_history.cell(row=row_idx, column=col_idx)
            header = headers_history[col_idx - 1]
            
            cell.border = thin_border
            
            if header in price_columns:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif '이격도' in str(header):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif '일' in str(header):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # History 헤더 포맷팅
    for col_idx in range(1, ws_history.max_column + 1):
        header_cell = ws_history.cell(row=1, column=col_idx)
        header_cell.border = thin_border
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

# 파일 저장
wb.save('output/trading_signals.xlsx')

print("=== 최종 완료 ===")
print("trading_signals.xlsx 업데이트 완료")
print("- NAVER와 한미반도체 알람상태 수정")
print("- 모든 가격 데이터에 천 단위 콤마 적용")
print("- Excel 포맷팅 적용")

# 5. 결과 확인
print(f"\n=== 결과 확인 ===")
naver = df_summary[df_summary['티커'] == '035420']
hanmi = df_summary[df_summary['티커'] == '042700']

if len(naver) > 0:
    n = naver.iloc[0]
    print(f"NAVER: {n['종목명']}")
    print(f"  알람상태: {n['알람상태']}")
    print(f"  상태메시지: {n['상태메시지']}")
    print(f"  종가: {n['종가']}원")
    print(f"  1차매수선: {n['1차매수선']}원")

if len(hanmi) > 0:
    h = hanmi.iloc[0]
    print(f"\n한미반도체: {h['종목명']}")
    print(f"  알람상태: {h['알람상태']}")
    print(f"  상태메시지: {h['상태메시지']}")
    print(f"  종가: {h['종가']}원")
    print(f"  1차매수선: {h['1차매수선']}원")




