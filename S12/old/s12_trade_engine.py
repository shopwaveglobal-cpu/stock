#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
S12 Leading Universe (1-year rolling)

요구사항 요약
- SUMMARY: 매일 거래대금 5,000억 이상(임계값 가변) 종목을 "축적식"으로 유지
  · 중복 종목은 한 줄만 유지하고, 5천억 초과 횟수 COUNT 누적
  · 종목의 "날짜"는 해당 종목이 처음 등장한 가장 과거 날짜
  · 최근 종가/저가, MA20(A), A의 -20%(B), B의 -10%(C), C의 -10%(D), STATUS
- DAILY : SUMMARY에 포함된 종목들의 최근 1년치 일봉(시/고/저/종) + A/B/C/D + STATUS
- 기간: "당일 기준 최근 1년" 자동 계산 (옵션으로 --from/--to 지정 가능)
- 데이터 소스: 키움 REST (랭킹: ka10032, 차트: ka10081).
  필요 시 기존 turnover_universe.xlsx로부터 부트스트랩도 지원(--universe-excel)

중요: 본 파일은 사용자의 명시 요청으로 APPKEY/SECRET을 코드에 하드코딩합니다.
보안 위험이 있으니, 외부 유출/저장소 커밋을 반드시 금지하세요.
"""

from __future__ import annotations
import os
import sys
import time
import argparse
import datetime as dt
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# -------------------------- 하드코딩된 키 (사용자 요청) --------------------------
APPKEY_DEFAULT  = "IweTdkYa8JWDUOa8NohVSVeOiJ1THDGd_2x050A8XcU"
SECRET_DEFAULT  = "eazu-jPNJpAsIVkaUTh3_88gUvXrCMJCwGF2AYRtBJs"

# -------------------------- 설정/상수 --------------------------
OUTPUT_XLSX    = "leading_universe.xlsx"
SHEET_SUMMARY  = "SUMMARY"
SHEET_DAILY    = "DAILY"

# Kiwoom REST
KIWOOM_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
KIWOOM_BASE_URL  = "https://api.kiwoom.com"
ENDPOINT_RANK    = "/api/dostk/rkinfo"   # ka10032
ENDPOINT_OHLC    = "/api/dostk/chart"    # ka10081
API_ID_RANK      = "ka10032"
API_ID_CHART     = "ka10081"

# 랭킹 고정 파라미터(전일 거래대금 TOP100 기준)
RANK_FIXED_BODY = {
    "qry_tp": "2",           # 2: 전일 거래대금 상위
    "rank_strt": "0",
    "rank_end":  "100",
    "stex_tp":   "3",       # 3: 통합 (KRX+KOSDAQ)
    "mang_stk_incls": "1",  # 관리종목 포함 (후단에서 필터)
}
MARKETS = ["001", "101"]     # 001: KRX, 101: KOSDAQ

# 제외 키워드 (ETF/ETN/인덱스형 등)
EXCLUDE_KEYWORDS = [
    "KODEX", "TIGER", "KBSTAR", "KOSEF", "ARIRANG", "HANARO", "SOL", "TREX", "ACE",
    "인버스", "레버리지", "선물", "ETF", "ETN",
]

# 계산 파라미터
MA_LOOKBACK = 20
BAND_A20    = 0.20  # B = A*(1-0.20)
DROP_10     = 0.10  # C = B*(1-0.10), D = C*(1-0.10)

# ------------------------- 유틸/엑셀 -------------------------
def ensure_book(path: str) -> None:
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_SUMMARY
        ws.append([
            "날짜","티커","종목명","COUNT_5k","최근종가","최근저가",
            "A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"
        ])
        wb.create_sheet(SHEET_DAILY)
        wb[SHEET_DAILY].append([
            "날짜","티커","종목명","open","high","low","close",
            "A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"
        ])
        wb.save(path)

def auto_fit_borders(path: str, sheet: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet]
    # width
    for ci, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            s = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 6
    # border
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value not in (None, ""):
                cell.border = border
    wb.save(path)

# ------------------------- 날짜 유틸 -------------------------
def default_window_1y(end: Optional[str] = None) -> Tuple[str, str]:
    if end is None:
        e = dt.date.today()
    else:
        e = dt.datetime.strptime(end, "%Y-%m-%d").date()
    s = e - dt.timedelta(days=365)
    return s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d")

# --------------------- 키움 세션/요청 ---------------------
@dataclass
class KiwoomSession:
    appkey: str
    secret: str
    token: Optional[str] = None

    def acquire(self) -> str:
        import requests
        headers = {"Content-Type": "application/json;charset=UTF-8"}
        body = {"grant_type": "client_credentials", "appkey": self.appkey, "secretkey": self.secret}
        r = requests.post(KIWOOM_TOKEN_URL, headers=headers, json=body, timeout=20)
        r.raise_for_status()
        data = r.json()
        self.token = data.get("token") or data.get("access_token")
        if not self.token:
            raise RuntimeError(f"token missing: {data}")
        return self.token

    def post(self, endpoint: str, api_id: str, body: dict, max_retry: int = 6) -> dict:
        import requests
        if not self.token:
            self.acquire()
        url = KIWOOM_BASE_URL + endpoint
        headers = {
            "authorization": f"Bearer {self.token}",
            "Content-Type": "application/json;charset=UTF-8",
            "api-id": api_id,
            "cont-yn": "N",
            "next-key": "",
        }
        last_err = None
        for attempt in range(max_retry):
            try:
                r = requests.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 429:
                    ra = r.headers.get("Retry-After")
                    sleep_s = float(ra) if (ra and str(ra).isdigit()) else (0.5 * (2 ** attempt))
                    time.sleep(sleep_s)
                    continue
                if 500 <= r.status_code < 600:
                    time.sleep(0.5 * (2 ** attempt))
                    continue
                r.raise_for_status()
                return r.json()
            except requests.RequestException as e:
                last_err = e
                time.sleep(0.5 * (2 ** attempt))
        if last_err:
            raise last_err
        raise RuntimeError("Unknown error")

# --------------------- 데이터 수집/정리 ---------------------
def fetch_rank_for_day(sess: KiwoomSession, ymd: str, threshold_eok: float, sleep: float) -> pd.DataFrame:
    """전일 거래대금 TOP100(통합) → 필터링(>= 임계값), ETF/ETN 제외, 억 단위 변환.
    반환 컬럼: 날짜, 티커, 종목명, 거래대금(억)
    - 응답 키 변동/빈 응답에 안전하게 동작하도록 가드 추가
    """
    def _first_key(d: dict, candidates: list, default=""):
        for k in candidates:
            if k in d and d.get(k) not in (None, ""):
                return d.get(k)
        return default

    def _sum_markets(market: str) -> pd.DataFrame:
        body = dict(RANK_FIXED_BODY)
        body.update({"mrkt_tp": market, "bas_dd": ymd.replace("-", "")})
        data = sess.post(ENDPOINT_RANK, API_ID_RANK, body)
        rows = data.get("rkinfo", [])
        if not isinstance(rows, list):
            rows = []
        recs = []
        for r in rows:
            tkr  = str(_first_key(r, ["stk_cd", "stck_shrn_iscd", "ISU_SRT_CD"])) .strip()
            name = str(_first_key(r, ["stk_kor_isnm", "stk_kor_nm", "ISU_ABBRV"])) .strip()
            amt100m = _first_key(r, ["tvol_tamt", "TOT_TR_PRC", "tot_tr_prc"], default="")
            try:
                amt100m = float(str(amt100m).replace(",", "")) if amt100m not in (None, "") else np.nan
            except Exception:
                amt100m = np.nan
            if tkr:
                recs.append({"티커": tkr, "종목명": name, "거래대금(억)": amt100m * 0.1})  # 100만원→억
        return pd.DataFrame(recs, columns=["티커","종목명","거래대금(억)"])

    # 두 시장 합산(동일 티커 합산)
    df_all: List[pd.DataFrame] = []
    for m in MARKETS:
        time.sleep(sleep)
        df_all.append(_sum_markets(m))
    if not any([not x.empty for x in df_all]):
        return pd.DataFrame(columns=["날짜","티커","종목명","거래대금(억)"])

    df = pd.concat(df_all, ignore_index=True)
    if df.empty:
        return pd.DataFrame(columns=["날짜","티커","종목명","거래대금(억)"])

    # 티커 정형화
    df["티커"] = df["티커"].astype(str).str.split("_").str[0]

    # 이름 필터(ETF/ETN 등 제외)
    ex = df["종목명"].astype(str)
    mask_ex = pd.Series(False, index=df.index)
    for kw in EXCLUDE_KEYWORDS:
        mask_ex = mask_ex | ex.str.contains(kw, na=False)
    df = df.loc[~mask_ex].copy()

    # 임계값 필터
    if "거래대금(억)" in df.columns:
        df = df.loc[df["거래대금(억)"] >= threshold_eok].copy()

    # 날짜 부여
    df.insert(0, "날짜", ymd)

    # 중복 티커 합산(혹시 남아 있다면)
    if not df.empty:
        df = df.groupby(["날짜","티커","종목명"], as_index=False)["거래대금(억)"].sum()
    return df

def fetch_ohlc_series(sess: KiwoomSession, ticker: str, start: str, end: str) -> pd.DataFrame:
    """일봉(ka10081) → DataFrame[date, open, high, low, close] (구간 필터/중복 제거/정렬 포함)"""
    ymd = lambda s: s.replace("-", "")
    body = {"stk_cd": ticker, "base_dt": ymd(start), "end_dt": ymd(end), "upd_stkpc_tp": "1"}
    data = sess.post(ENDPOINT_OHLC, API_ID_CHART, body)
    rows = data.get("stk_dt_pole_chart_qry", [])
    if not isinstance(rows, list):
        rows = []
    def _f(x):
        if x in (None, ""): return np.nan
        try: return float(str(x).replace(",", ""))
        except Exception: return np.nan
    parsed = []
    for r in rows:
        dd = str(r.get("dt", "")).strip()
        op, hi, lo, cl = _f(r.get("open_pric")), _f(r.get("high_pric")), _f(r.get("low_pric")), _f(r.get("cur_prc"))
        if not dd or np.isnan(cl):
            continue
        dd = dd.replace(".", "").replace("-", "")
        dd = f"{dd[:4]}-{dd[4:6]}-{dd[6:]}" if len(dd) == 8 else dd
        parsed.append({"date": dd, "open": op, "high": hi, "low": lo, "close": cl})
    df = pd.DataFrame(parsed)
    if df.empty:
        return df
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df = df.dropna(subset=["date"]).copy()
    s = pd.to_datetime(start).date(); e = pd.to_datetime(end).date()
    df = df[(df["date"] >= s) & (df["date"] <= e)].copy()
    df = df.drop_duplicates(subset=["date"]).sort_values("date").reset_index(drop=True)
    return df

# --------------------- 계산(A/B/C/D, 상태) ---------------------
def add_bands(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["A(MA20)"] = out["close"].rolling(MA_LOOKBACK).mean()
    out["B(A*0.8)"] = out["A(MA20)"] * (1 - BAND_A20)
    out["C(B*0.9)"] = out["B(A*0.8)"] * (1 - DROP_10)
    out["D(C*0.9)"] = out["C(B*0.9)"] * (1 - DROP_10)
    return out

def infer_status_row(last_close: float, last_B: float) -> str:
    try:
        if np.isnan(last_close) or np.isnan(last_B):
            return "NA"
        if last_close <= last_B:
            return "BUY_ZONE"
        return "WATCH"
    except Exception:
        return "NA"

# --------------------- SUMMARY/DAILY 업데이트 ---------------------
def update_summary(sess: KiwoomSession, out_xlsx: str, start: str, end: str, threshold_eok: float, use_api: int, universe_excel: Optional[str], sleep: float) -> pd.DataFrame:
    """1) (선택) API에서 1년 구간의 일별 랭킹을 돌며 주도주 축적
       2) 또는 기존 universe 엑셀에서 1년 구간 필터로 부트스트랩
       3) SUMMARY 갱신/COUNT 누적/최신가 지표 업데이트
    반환: SUMMARY DataFrame
    """
    ensure_book(out_xlsx)
    try:
        df_sum = pd.read_excel(out_xlsx, sheet_name=SHEET_SUMMARY, dtype={"티커": str})
    except Exception:
        df_sum = pd.DataFrame(columns=[
            "날짜","티커","종목명","COUNT_5k","최근종가","최근저가","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"
        ])

    # 1) 소스 유니버스 구축
    if use_api:
        # 영업일 리스트(주말 제외)
        days = pd.bdate_range(start=start, end=end).strftime("%Y-%m-%d").tolist()
        uni_rows = []
        for ymd in days:
            try:
                time.sleep(sleep)
                df = fetch_rank_for_day(sess, ymd, threshold_eok, sleep)
                if not df.empty:
                    uni_rows.append(df)
            except Exception as e:
                print(f"[WARN] rank {ymd} failed: {e}")
                continue
        uni = pd.concat(uni_rows, ignore_index=True) if uni_rows else pd.DataFrame(columns=["날짜","티커","종목명","거래대금(억)"])
    else:
        # turnover_universe.xlsx 로부터 1년치 필터
        if not universe_excel or not os.path.exists(universe_excel):
            raise RuntimeError("universe 엑셀이 없거나 경로가 잘못되었습니다. --use-api 1 로 직접 수집을 선택하세요.")
        uni = pd.read_excel(universe_excel, sheet_name="universe", dtype={"티커": str})
        uni["날짜"] = pd.to_datetime(uni["날짜"], errors="coerce").dt.date
        s = pd.to_datetime(start).date(); e = pd.to_datetime(end).date()
        uni = uni[(uni["날짜"] >= s) & (uni["날짜"] <= e) & (uni["거래대금(억)"] >= threshold_eok)].copy()

    # 2) SUMMARY 축적/COUNT 누적/최초 날짜
    if not uni.empty:
        first_date = (uni.groupby(["티커","종목명"]) ["날짜"].min()).reset_index().rename(columns={"날짜":"_first"})
        count_5k  = (uni.groupby(["티커"]) ["날짜"].count()).reset_index().rename(columns={"날짜":"_count"})
        base = first_date.merge(count_5k, on="티커", how="left")
        base["_count"] = base["_count"].fillna(0).astype(int)

        if not df_sum.empty:
            df_sum["날짜"] = pd.to_datetime(df_sum["날짜"], errors="coerce").dt.date
            merged = df_sum.merge(base[["티커","_first","_count"]], on="티커", how="outer")
            # 날짜: 둘 중 더 과거
            merged["날짜"] = merged.apply(lambda r: (min([d for d in [r.get("날짜"), r.get("_first")] if pd.notna(d)]) if (pd.notna(r.get("날짜")) or pd.notna(r.get("_first"))) else pd.NaT), axis=1)
            # COUNT 누적
            merged["COUNT_5k"] = merged.get("COUNT_5k", 0).fillna(0).astype(int) + merged.get("_count", 0).fillna(0).astype(int)
            # 종목명 보정
            merged["종목명"] = merged["종목명"].fillna("")
            for _, rr in base.iterrows():
                t = rr["티커"]
                if (merged.loc[merged["티커"]==t, "종목명"]=="").any():
                    name = uni.loc[uni["티커"]==t, "종목명"].dropna().astype(str)
                    if not name.empty:
                        merged.loc[merged["티커"]==t, "종목명"] = name.iloc[0]
            df_sum = merged
        else:
            # 신규 생성
            df_sum = base.rename(columns={"_first":"날짜"}).copy()
            df_sum.insert(2, "종목명", df_sum["티커"].map(lambda t: uni.loc[uni["티커"]==t, "종목명"].dropna().astype(str).iloc[0] if not uni.loc[uni["티커"]==t, "종목명"].dropna().empty else ""))
            df_sum = df_sum.rename(columns={"_count":"COUNT_5k"})
            df_sum["최근종가"] = np.nan
            df_sum["최근저가"] = np.nan
            df_sum["A(MA20)"] = np.nan
            df_sum["B(A*0.8)"] = np.nan
            df_sum["C(B*0.9)"] = np.nan
            df_sum["D(C*0.9)"] = np.nan
            df_sum["STATUS"]  = "NA"

    # 3) 최근가/지표 업데이트
    if not df_sum.empty:
        s, e = start, end
        rows_update = []
        for i, r in df_sum.iterrows():
            tkr = str(r["티커"]).zfill(6)
            name = str(r.get("종목명", "")).strip()
            try:
                time.sleep(sleep)
                ohlc = fetch_ohlc_series(sess, tkr, s, e)
                if ohlc.empty:
                    continue
                env = add_bands(ohlc)
                last = env.dropna(subset=["close"]).iloc[-1]
                last_close = float(last["close"]) if pd.notna(last["close"]) else np.nan
                last_low   = float(last["low"]) if pd.notna(last["low"]) else np.nan
                last_A     = float(last["A(MA20)"]) if pd.notna(last["A(MA20)"]) else np.nan
                last_B     = float(last["B(A*0.8)"]) if pd.notna(last["B(A*0.8)"]) else np.nan
                last_C     = float(last["C(B*0.9)"]) if pd.notna(last["C(B*0.9)"]) else np.nan
                last_D     = float(last["D(C*0.9)"]) if pd.notna(last["D(C*0.9)"]) else np.nan
                status     = infer_status_row(last_close, last_B)
                rows_update.append((i, last_close, last_low, last_A, last_B, last_C, last_D, status))
            except Exception as e:
                print(f"[WARN] SUMMARY update {tkr}({name}) failed: {e}")
                continue
        for (idx, lc, ll, a, b, c, d, st) in rows_update:
            df_sum.at[idx, "최근종가"] = lc
            df_sum.at[idx, "최근저가"] = ll
            df_sum.at[idx, "A(MA20)"] = a
            df_sum.at[idx, "B(A*0.8)"] = b
            df_sum.at[idx, "C(B*0.9)"] = c
            df_sum.at[idx, "D(C*0.9)"] = d
            df_sum.at[idx, "STATUS"]   = st

    # 저장 (SUMMARY만 우선 기록, DAILY는 update_daily에서 함께 기록)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df_sum.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
    auto_fit_borders(OUTPUT_XLSX, SHEET_SUMMARY)
    return df_sum


def update_daily(sess: KiwoomSession, out_xlsx: str, summary_df: pd.DataFrame, start: str, end: str, sleep: float) -> pd.DataFrame:
    ensure_book(out_xlsx)
    try:
        df_daily = pd.read_excel(out_xlsx, sheet_name=SHEET_DAILY, dtype={"티커": str})
    except Exception:
        df_daily = pd.DataFrame(columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"])

    s, e = start, end
    new_rows: List[List[Any]] = []
    for _, r in summary_df.iterrows():
        tkr = str(r["티커"]).zfill(6)
        name = str(r.get("종목명", "")).strip()
        try:
            time.sleep(sleep)
            ohlc = fetch_ohlc_series(sess, tkr, s, e)
            if ohlc.empty:
                continue
            env = add_bands(ohlc)
            env["STATUS"] = env.apply(lambda x: infer_status_row(x.get("close", np.nan), x.get("B(A*0.8)", np.nan)), axis=1)
            dd = env[["date","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]].copy()
            dd.insert(1, "티커", tkr)
            dd.insert(2, "종목명", name)
            dd.rename(columns={"date":"날짜"}, inplace=True)
            # 기존과 중복 제거 후만 추가
            if not df_daily.empty:
                cur = df_daily.loc[df_daily["티커"]==tkr, ["날짜"]]
                dd = dd.merge(cur, on=["날짜"], how="left", indicator=True)
                dd = dd[dd["_merge"]=="left_only"].drop(columns=["_merge"])  # 신규만
            new_rows.extend(dd.values.tolist())
        except Exception as e:
            print(f"[WARN] DAILY update {tkr}({name}) failed: {e}")
            continue

    if new_rows:
        df_new = pd.DataFrame(new_rows, columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"])
        df_daily = pd.concat([df_daily, df_new], ignore_index=True)
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            # SUMMARY 다시 기록
            summary_df.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
            df_daily.to_excel(writer, index=False, sheet_name=SHEET_DAILY)
        auto_fit_borders(out_xlsx, SHEET_SUMMARY)
        auto_fit_borders(out_xlsx, SHEET_DAILY)
    return df_daily

# ------------------------------ CLI ------------------------------
def main():
    ap = argparse.ArgumentParser(description="S12 Leading Universe (1-year rolling)")
    ap.add_argument("--from", dest="from_date", default=None, help="YYYY-MM-DD (default: today-365d)")
    ap.add_argument("--to", dest="to_date", default=None, help="YYYY-MM-DD (default: today)")
    ap.add_argument("--threshold-eok", type=float, default=5000.0, help="거래대금 임계값(억), default 5000")
    ap.add_argument("--out", default=OUTPUT_XLSX, help="출력 엑셀 경로")
    ap.add_argument("--use-api", type=int, default=1, help="1: 랭킹 API로 1년치 수집, 0: --universe-excel 사용")
    ap.add_argument("--universe-excel", default=None, help="turnover_universe.xlsx 경로 (use-api=0일 때만)")
    ap.add_argument("--sleep", type=float, default=0.25, help="API 호출 간 대기(초)")
    ap.add_argument("--appkey", default=None, help="키움 APPKEY (환경변수/하드코딩 우선순위는: 인자 > 환경변수 > 하드코딩)")
    ap.add_argument("--secret", default=None, help="키움 SECRET (환경변수/하드코딩 우선순위는: 인자 > 환경변수 > 하드코딩)")
    args = ap.parse_args()

    if args.from_date and args.to_date:
        start, end = args.from_date, args.to_date
    else:
        start, end = default_window_1y(args.to_date)

    # 우선순위: 인자 > 환경변수 > 하드코딩
    appkey = args.appkey or os.getenv("KIWOOM_APPKEY") or APPKEY_DEFAULT
    secret = args.secret or os.getenv("KIWOOM_SECRET") or SECRET_DEFAULT

    sess = KiwoomSession(appkey, secret)

    # SUMMARY 업데이트
    df_sum = update_summary(sess, args.out, start, end, args.threshold_eok, args.use_api, args.universe_excel, args.sleep)
    # DAILY 업데이트
    update_daily(sess, args.out, df_sum, start, end, args.sleep)
    print(f"[DONE] window={start}~{end} → {args.out}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Fatal: {e}")
        sys.exit(1)
