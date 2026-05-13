#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
S12 Leading Universe (1-year rolling)

요구사항(최종)
- SUMMARY: "당일 기준 최근 1년" 구간에서 거래대금 5,000억 이상(임계값 가변) 종목을 축적식으로 저장
  · 중복 종목은 1행만 유지, "COUNT_5k"는 해당 조건을 만족한 **영업일 수(고유 날짜수)** 로 누적
  · "날짜"는 해당 종목의 **가장 과거(최초 등장) 날짜**
  · 최근 종가/저가, A=MA20, B=A*0.8, C=B*0.9, D=C*0.9, STATUS 저장
- DAILY: SUMMARY 종목의 최근 1년치 일봉(시/고/저/종) + A/B/C/D + STATUS (중복 일자 스킵)
- 티커 정규화: '011200_AL' 등 접미사는 **모두 제거**하여 **6자리 숫자**로 통일(동일 종목으로 집계)
- 기간: 기본은 오늘 기준 1년. `--from/--to`로 명시적 지정 가능
- 소스: (A) 키움 랭킹 API (최근일만 신뢰) 또는 (B) 기존 `turnover_universe.xlsx`의 `universe` 시트(권장 backfill)
- 키: 사용자 요청에 따라 **코드에 하드코딩** (외부 유출 금지)

사용 예시
  기본(오늘 기준 1년):
    python -u s12_leading_universe.py

  기존 universe 파일로 1년 롤링 구성(권장):
    python -u s12_leading_universe.py --use-api 0 --universe-excel turnover_universe.xlsx

  기간 지정 + 임계값 변경:
    python -u s12_leading_universe.py --from 2025-09-01 --to 2025-10-08 --threshold-eok 5000 --use-api 0 --universe-excel turnover_universe.xlsx
"""

from __future__ import annotations
import os
import sys
import time
import argparse
import datetime as dt
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# -------------------------- 하드코딩된 키(사용자 요청) --------------------------
APPKEY_DEFAULT = "IweTdkYa8JWDUOa8NohVSVeOiJ1THDGd_2x050A8XcU"
SECRET_DEFAULT = "eazu-jPNJpAsIVkaUTh3_88gUvXrCMJCwGF2AYRtBJs"

# -------------------------- 설정/상수 --------------------------
OUTPUT_XLSX   = "leading_universe.xlsx"
SHEET_SUMMARY = "SUMMARY"
SHEET_DAILY   = "DAILY"

# Kiwoom REST
KIWOOM_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
KIWOOM_BASE_URL  = "https://api.kiwoom.com"
ENDPOINT_RANK    = "/api/dostk/rkinfo"   # ka10032
ENDPOINT_OHLC    = "/api/dostk/chart"    # ka10081
API_ID_RANK      = "ka10032"
API_ID_CHART     = "ka10081"

# 랭킹 고정 파라미터(전일 거래대금 TOP100)
RANK_FIXED_BODY = {
    "qry_tp": "2",           # 2: 전일 거래대금 상위
    "rank_strt": "0",
    "rank_end":  "100",
    "stex_tp":   "3",       # 3: 통합 (KRX+KOSDAQ)
    "mang_stk_incls": "1",  # 관리종목 포함(후단에서 필터)
}
MARKETS = ["001", "101"]     # 001: 유가, 101: 코스닥

# 제외 키워드(ETF/ETN/지수형 등)
EXCLUDE_KEYWORDS = [
    "KODEX", "TIGER", "KBSTAR", "KOSEF", "ARIRANG", "HANARO", "SOL", "TREX", "ACE",
    "인버스", "레버리지", "선물", "ETF", "ETN",
]

# 계산 파라미터
MA_LOOKBACK = 20
BAND_A20    = 0.20  # B = A*(1-0.20)
DROP_10     = 0.10  # C = B*(1-0.10), D = C*(1-0.10)

# ------------------------- 엑셀 유틸 -------------------------
def ensure_book(path: str) -> None:
    if not os.path.exists(path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = SHEET_SUMMARY
        ws1.append([
            "날짜","티커","종목명","COUNT_5k","최근종가","최근저가",
            "A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"
        ])
        ws2 = wb.create_sheet(SHEET_DAILY)
        ws2.append([
            "날짜","티커","종목명","open","high","low","close",
            "A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"
        ])
        wb.save(path)

def auto_fit_borders(path: str, sheet: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet]
    # width
    for ci, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            s = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 6
    # border
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value not in (None, ""):
                cell.border = border
    wb.save(path)

# ------------------------- 날짜/티커 유틸 -------------------------
def default_window_1y(end: Optional[str] = None) -> Tuple[str, str]:
    if end is None:
        e = dt.date.today()
    else:
        e = dt.datetime.strptime(end, "%Y-%m-%d").date()
    s = e - dt.timedelta(days=365)
    return s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d")

def normalize_ticker(t: Any) -> str:
    """011200_AL → 011200, '11200' → 011200, None → ''"""
    if t is None:
        return ""
    s = str(t).strip()
    if "_" in s:
        s = s.split("_")[0]
    s = ''.join(ch for ch in s if ch.isdigit())
    return s.zfill(6)[:6] if s else ""

# --------------------- 키움 세션/요청 ---------------------
@dataclass
class KiwoomSession:
    appkey: str
    secret: str
    token: Optional[str] = None

    def acquire(self) -> str:
        import requests
        headers = {"Content-Type": "application/json;charset=UTF-8"}
        body = {"grant_type": "client_credentials", "appkey": self.appkey, "secretkey": self.secret}
        r = requests.post(KIWOOM_TOKEN_URL, headers=headers, json=body, timeout=20)
        r.raise_for_status()
        data = r.json()
        self.token = data.get("token") or data.get("access_token")
        if not self.token:
            raise RuntimeError(f"token missing: {data}")
        return self.token

    def post(self, endpoint: str, api_id: str, body: dict, max_retry: int = 6) -> dict:
        import requests
        if not self.token:
            self.acquire()
        url = KIWOOM_BASE_URL + endpoint
        headers = {
            "authorization": f"Bearer {self.token}",
            "Content-Type": "application/json;charset=UTF-8",
            "api-id": api_id,
            "cont-yn": "N",
            "next-key": "",
        }
        last_err = None
        for attempt in range(max_retry):
            try:
                r = requests.post(url, headers=headers, json=body, timeout=20)
                if r.status_code == 429:
                    time.sleep(0.5 * (2 ** attempt))
                    continue
                if 500 <= r.status_code < 600:
                    time.sleep(0.5 * (2 ** attempt))
                    continue
                r.raise_for_status()
                return r.json()
            except requests.RequestException as e:
                last_err = e
                time.sleep(0.5 * (2 ** attempt))
        if last_err:
            raise last_err
        raise RuntimeError("Unknown error")

# --------------------- 데이터 수집 ---------------------
def fetch_rank_for_day(sess: KiwoomSession, ymd: str, threshold_eok: float, sleep: float) -> pd.DataFrame:
    """전일 거래대금 TOP100(통합) → 필터링(>= 임계값), ETF/ETN 제외, 억 단위 변환.
    반환 컬럼: 날짜, 티커, 종목명, 거래대금(억)
    응답 키 변동/빈 응답에 안전하게 동작.
    """
    def _first_key(d: dict, candidates: list, default=""):
        for k in candidates:
            if k in d and d.get(k) not in (None, ""):
                return d.get(k)
        return default

    def _sum_market(market: str) -> pd.DataFrame:
        body = dict(RANK_FIXED_BODY)
        body.update({"mrkt_tp": market, "bas_dd": ymd.replace("-", "")})
        data = sess.post(ENDPOINT_RANK, API_ID_RANK, body)
        rows = data.get("rkinfo", [])
        if not isinstance(rows, list):
            rows = []
        recs = []
        for r in rows:
            raw_tkr = _first_key(r, ["stk_cd", "stck_shrn_iscd", "ISU_SRT_CD"])  # 후보 키
            tkr     = normalize_ticker(raw_tkr)
            name    = str(_first_key(r, ["stk_kor_isnm", "stk_kor_nm", "ISU_ABBRV"])) .strip()
            amt100m = _first_key(r, ["tvol_tamt", "TOT_TR_PRC", "tot_tr_prc"], default="")
            try:
                amt100m = float(str(amt100m).replace(",", "")) if amt100m not in (None, "") else np.nan
            except Exception:
                amt100m = np.nan
            if not tkr:
                continue
            if any(kw in name for kw in EXCLUDE_KEYWORDS):
                continue
            recs.append({"날짜": ymd, "티커": tkr, "종목명": name, "거래대금(억)": amt100m * 0.1})  # 100만원→억
        return pd.DataFrame(recs, columns=["날짜","티커","종목명","거래대금(억)"])

    dfs = []
    for m in MARKETS:
        time.sleep(sleep)
        dfs.append(_sum_market(m))
    df = pd.concat(dfs, ignore_index=True) if any([not d.empty for d in dfs]) else pd.DataFrame(columns=["날짜","티커","종목명","거래대금(억)"])

    # 임계값 필터 + 날짜/티커 중복 제거(하루에 시장 2곳 중복 방지)
    if not df.empty:
        df = df[df["거래대금(억)"].astype(float) >= threshold_eok].copy()
        df = df.drop_duplicates(subset=["날짜","티커"])  # 고유 날짜-티커만 유지
    return df


def fetch_ohlc_series(sess: KiwoomSession, ticker: str, start: str, end: str) -> pd.DataFrame:
    """일봉(ka10081) → DataFrame[date, open, high, low, close]
    - 다양한 응답 포맷/키를 허용(서버 버전/문서와 실제 차이 대응)
    - 구간 필터/중복 제거/정렬 포함
    """
    def _first_key(d: dict, cands: list, default=None):
        for k in cands:
            if k in d and d.get(k) not in (None, ""):
                return d.get(k)
        return default

    ymd = lambda s: s.replace("-", "")
    body = {"stk_cd": ticker, "base_dt": ymd(start), "end_dt": ymd(end), "upd_stkpc_tp": "1"}
    data = sess.post(ENDPOINT_OHLC, API_ID_CHART, body)

    # 응답 컨테이너 후보
    rows = (
        data.get("stk_dt_pole_chart_qry") or
        data.get("output") or
        data.get("output2") or
        data.get("chart") or
        []
    )
    if not isinstance(rows, list):
        rows = []

    def _to_float(x):
        if x in (None, ""): return np.nan
        try:
            return float(str(x).replace(",", ""))
        except Exception:
            return np.nan

    parsed = []
    for r in rows:
        # 날짜 후보
        dd = _first_key(r, ["dt", "stck_bsop_date", "date", "bas_dt"], default="")
        if not dd:
            continue
        dd = str(dd).strip().replace(".", "").replace("-", "")
        if len(dd) == 8:
            dd = f"{dd[:4]}-{dd[4:6]}-{dd[6:]}"

        # 가격 키 후보 (키움/표준 혼합 대응)
        op = _first_key(r, ["open_pric", "stck_oprc", "open"], default="")
        hi = _first_key(r, ["high_pric", "stck_hgpr", "high"], default="")
        lo = _first_key(r, ["low_pric",  "stck_lwpr", "low" ], default="")
        cl = _first_key(r, ["cur_prc",   "stck_clpr", "close"], default="")
        op, hi, lo, cl = _to_float(op), _to_float(hi), _to_float(lo), _to_float(cl)
        if np.isnan(cl):
            continue
        parsed.append({"date": dd, "open": op, "high": hi, "low": lo, "close": cl})

    df = pd.DataFrame(parsed)
    if df.empty:
        return df
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df = df.dropna(subset=["date"]).copy()
    s = pd.to_datetime(start).date(); e = pd.to_datetime(end).date()
    df = df[(df["date"] >= s) & (df["date"] <= e)].copy()
    df = df.drop_duplicates(subset=["date"]).sort_values("date").reset_index(drop=True)
    return df
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df = df.dropna(subset=["date"]).copy()
    s = pd.to_datetime(start).date(); e = pd.to_datetime(end).date()
    df = df[(df["date"] >= s) & (df["date"] <= e)].copy()
    df = df.drop_duplicates(subset=["date"]).sort_values("date").reset_index(drop=True)
    return df

# --------------------- 계산(A/B/C/D, 상태) ---------------------
def add_bands(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["A(MA20)"] = out["close"].rolling(MA_LOOKBACK, min_periods=1).mean()
    out["B(A*0.8)"] = out["A(MA20)"] * (1 - BAND_A20)
    out["C(B*0.9)"] = out["B(A*0.8)"] * (1 - DROP_10)
    out["D(C*0.9)"] = out["C(B*0.9)"] * (1 - DROP_10)
    return out

def infer_status_row(last_close: float, last_B: float) -> str:
    try:
        if np.isnan(last_close) or np.isnan(last_B):
            return "NA"
        return "BUY_ZONE" if last_close <= last_B else "WATCH"
    except Exception:
        return "NA"

# --------------------- SUMMARY/DAILY 업데이트 ---------------------
def update_summary(sess: KiwoomSession, out_xlsx: str, start: str, end: str, threshold_eok: float, use_api: int, universe_excel: Optional[str], sleep: float) -> pd.DataFrame:
    ensure_book(out_xlsx)
    try:
        df_sum = pd.read_excel(out_xlsx, sheet_name=SHEET_SUMMARY, dtype={"티커": str})
    except Exception:
        df_sum = pd.DataFrame(columns=["날짜","티커","종목명","COUNT_5k","최근종가","최근저가","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"])

    # 1) 유니버스 구축
    if use_api:
        days = pd.bdate_range(start=start, end=end).strftime("%Y-%m-%d").tolist()
        print(f"[SUMMARY] collecting rank days: {len(days)} ({start}~{end})", flush=True)
        uni_rows = []
        for i, ymd in enumerate(days, start=1):
            print(f"  - [{i}/{len(days)}] rank {ymd} ...", flush=True)
            try:
                time.sleep(sleep)
                df = fetch_rank_for_day(sess, ymd, threshold_eok, sleep)
                if not df.empty:
                    uni_rows.append(df)
            except Exception as e:
                print(f"[WARN] rank {ymd} failed: {e}", flush=True)
                continue
        uni = pd.concat(uni_rows, ignore_index=True) if uni_rows else pd.DataFrame(columns=["날짜","티커","종목명","거래대금(억)"])
        print(f"[SUMMARY] rank rows collected: {0 if uni_rows==[] else len(uni)}", flush=True)
    else:
        if not universe_excel or not os.path.exists(universe_excel):
            raise RuntimeError("universe 엑셀이 없거나 경로가 잘못되었습니다. --use-api 1 로 직접 수집을 선택하세요.")
        uni = pd.read_excel(universe_excel, sheet_name="universe", dtype={"티커": str})
        # 정규화
        uni["티커"] = uni["티커"].map(normalize_ticker)
        uni["날짜"] = pd.to_datetime(uni["날짜"], errors="coerce").dt.date
        s = pd.to_datetime(start).date(); e = pd.to_datetime(end).date()
        uni = uni[(uni["날짜"] >= s) & (uni["날짜"] <= e) & (uni["거래대금(억)"] >= threshold_eok)].copy()
        # 하루에 중복된 시장행 제거 → 고유 날짜-티커만 유지
        uni = uni.drop_duplicates(subset=["날짜","티커"]).copy()
        print(f"[SUMMARY] bootstrap rows(after dedup): {len(uni)}", flush=True)

    if uni.empty:
        print("[SUMMARY] universe empty → 이후 단계는 스킵될 수 있습니다.", flush=True)
        # SUMMARY/DAILY 템플릿만 유지 저장
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            df_sum.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
            pd.DataFrame(columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]).to_excel(writer, index=False, sheet_name=SHEET_DAILY)
        auto_fit_borders(out_xlsx, SHEET_SUMMARY)
        auto_fit_borders(out_xlsx, SHEET_DAILY)
        return df_sum

    # 2) SUMMARY 베이스: 최초 날짜 + COUNT(고유 날짜수)
    first_date = (uni.groupby(["티커","종목명"])["날짜"].min().reset_index().rename(columns={"날짜":"_first"}))
    # COUNT는 고유 날짜 수로 계산
    count_5k  = (uni.groupby(["티커"])["날짜"].nunique().reset_index().rename(columns={"날짜":"_count"}))
    base = first_date.merge(count_5k, on="티커", how="left")
    base["_count"] = base["_count"].fillna(0).astype(int)

    # 기존 SUMMARY와 병합(축적)
    if not df_sum.empty:
        df_sum["날짜"] = pd.to_datetime(df_sum["날짜"], errors="coerce").dt.date
        merged = df_sum.merge(base[["티커","종목명","_first","_count"]], on="티커", how="outer", suffixes=("","_new"))
        # 종목명: 신규가 있으면 보정
        merged["종목명"] = merged.apply(lambda r: r["종목명"] if isinstance(r["종목명"], str) and r["종목명"].strip() else r.get("종목명_new", r.get("종목명")), axis=1)
        # 날짜: 더 과거값 유지
        def _min_date(a, b):
            da = a if pd.notna(a) else None
            db = b if pd.notna(b) else None
            if da and db:
                return min(da, db)
            return da or db
        merged["날짜"] = merged.apply(lambda r: _min_date(r.get("날짜"), r.get("_first")), axis=1)
        # COUNT 누적(이중계산 방지: 신규 _count만 더함)
        merged["COUNT_5k"] = merged.get("COUNT_5k", 0).fillna(0).astype(int) + merged.get("_count", 0).fillna(0).astype(int)
        df_sum = merged[["날짜","티커","종목명","COUNT_5k","최근종가","최근저가","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]].copy()
    else:
        df_sum = base.rename(columns={"_first":"날짜","_count":"COUNT_5k"}).copy()
        # 종목명 컬럼 안전 삽입/갱신
        name_map = df_sum["티커"].map(lambda t: uni.loc[uni["티커"]==t, "종목명"].dropna().astype(str).iloc[0] if not uni.loc[uni["티커"]==t, "종목명"].dropna().empty else "")
        if "종목명" in df_sum.columns:
            df_sum["종목명"] = name_map
        else:
            df_sum.insert(2, "종목명", name_map)
        # 지표 컬럼 초기화
        for col in ["최근종가","최근저가","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]:
            df_sum[col] = np.nan
        df_sum["STATUS"] = "NA"

    # 3) 최신가/지표 계산(요청이 많으니 슬립)
    s, e = start, end
    rows_update = []
    for i, r in df_sum.iterrows():
        tkr = normalize_ticker(r["티커"])  # 안전
        name = str(r.get("종목명", "")).strip()
        try:
            time.sleep(sleep)
            ohlc = fetch_ohlc_series(sess, tkr, s, e)
            if ohlc.empty:
                continue
            env = add_bands(ohlc)
            last = env.dropna(subset=["close"]).iloc[-1]
            last_close = float(last["close"]) if pd.notna(last["close"]) else np.nan
            last_low   = float(last["low"]) if pd.notna(last["low"]) else np.nan
            last_A     = float(last["A(MA20)"]) if pd.notna(last["A(MA20)"]) else np.nan
            last_B     = float(last["B(A*0.8)"]) if pd.notna(last["B(A*0.8)"]) else np.nan
            last_C     = float(last["C(B*0.9)"]) if pd.notna(last["C(B*0.9)"]) else np.nan
            last_D     = float(last["D(C*0.9)"]) if pd.notna(last["D(C*0.9)"]) else np.nan
            status     = infer_status_row(last_close, last_B)
            rows_update.append((i, last_close, last_low, last_A, last_B, last_C, last_D, status))
        except Exception as ex:
            print(f"[WARN] SUMMARY update {tkr}({name}) failed: {ex}")
            continue
    for (idx, lc, ll, a, b, c, d, st) in rows_update:
        df_sum.at[idx, "최근종가"] = lc
        df_sum.at[idx, "최근저가"] = ll
        df_sum.at[idx, "A(MA20)"] = a
        df_sum.at[idx, "B(A*0.8)"] = b
        df_sum.at[idx, "C(B*0.9)"] = c
        df_sum.at[idx, "D(C*0.9)"] = d
        df_sum.at[idx, "STATUS"]   = st

    # 저장: SUMMARY와 DAILY 템플릿 동시 유지
    try:
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            df_sum.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
            pd.DataFrame(columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]).to_excel(writer, index=False, sheet_name=SHEET_DAILY)
    except PermissionError:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = out_xlsx.replace(".xlsx", f"_{ts}.xlsx")
        with pd.ExcelWriter(alt, engine="openpyxl") as writer:
            df_sum.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
            pd.DataFrame(columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]).to_excel(writer, index=False, sheet_name=SHEET_DAILY)
        print(f"[DONE] SUMMARY saved as {alt}")
    auto_fit_borders(out_xlsx, SHEET_SUMMARY)
    auto_fit_borders(out_xlsx, SHEET_DAILY)
    print(f"[SUMMARY] symbols: {len(df_sum)}")
    return df_sum


def update_daily(sess: KiwoomSession, out_xlsx: str, summary_df: pd.DataFrame, start: str, end: str, sleep: float) -> pd.DataFrame:
    ensure_book(out_xlsx)
    try:
        df_daily = pd.read_excel(out_xlsx, sheet_name=SHEET_DAILY, dtype={"티커": str})
    except Exception:
        df_daily = pd.DataFrame(columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"])

    s, e = start, end
    new_rows: List[List[Any]] = []
    miss_cnt = 0
    ok_cnt = 0

    for _, r in summary_df.iterrows():
        tkr = normalize_ticker(r["티커"])
        name = str(r.get("종목명", "")).strip()
        try:
            time.sleep(sleep)
            ohlc = fetch_ohlc_series(sess, tkr, s, e)
            if ohlc.empty:
                miss_cnt += 1
                print(f"[DAILY] no OHLC for {tkr}({name})", flush=True)
                continue
            ok_cnt += 1
            env = add_bands(ohlc)
            env["STATUS"] = env.apply(lambda x: infer_status_row(x.get("close", np.nan), x.get("B(A*0.8)", np.nan)), axis=1)
            dd = env[["date","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"]].copy()
            dd.insert(1, "티커", tkr)
            dd.insert(2, "종목명", name)
            dd.rename(columns={"date":"날짜"}, inplace=True)
            if not df_daily.empty:
                cur = df_daily.loc[df_daily["티커"]==tkr, ["날짜"]]
                dd = dd.merge(cur, on=["날짜"], how="left", indicator=True)
                dd = dd[dd["_merge"]=="left_only"].drop(columns=["_merge"])  # 신규만
            if not dd.empty:
                new_rows.extend(dd.values.tolist())
        except Exception as ex:
            print(f"[WARN] DAILY update {tkr}({name}) failed: {ex}", flush=True)
            continue

    print(f"[DAILY] OHLC ok={ok_cnt}, miss={miss_cnt}", flush=True)

    if new_rows:
        df_new = pd.DataFrame(new_rows, columns=["날짜","티커","종목명","open","high","low","close","A(MA20)","B(A*0.8)","C(B*0.9)","D(C*0.9)","STATUS"])
        df_daily = pd.concat([df_daily, df_new], ignore_index=True)
        try:
            with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
                summary_df.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
                df_daily.to_excel(writer, index=False, sheet_name=SHEET_DAILY)
        except PermissionError:
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            alt = out_xlsx.replace(".xlsx", f"_{ts}.xlsx")
            with pd.ExcelWriter(alt, engine="openpyxl") as writer:
                summary_df.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY)
                df_daily.to_excel(writer, index=False, sheet_name=SHEET_DAILY)
            print(f"[DONE] file locked, saved as → {alt}")
        auto_fit_borders(out_xlsx, SHEET_SUMMARY)
        auto_fit_borders(out_xlsx, SHEET_DAILY)
    else:
        print("[DAILY] 신규 추가 행 없음", flush=True)

    return df_daily

# ------------------------------ CLI ------------------------------
def main():
    ap = argparse.ArgumentParser(description="S12 Leading Universe (1-year rolling)")
    ap.add_argument("--from", dest="from_date", default=None, help="YYYY-MM-DD (default: today-365d)")
    ap.add_argument("--to", dest="to_date", default=None, help="YYYY-MM-DD (default: today)")
    ap.add_argument("--threshold-eok", type=float, default=5000.0, help="거래대금 임계값(억), default 5000")
    ap.add_argument("--out", default=OUTPUT_XLSX, help="출력 엑셀 경로")
    ap.add_argument("--use-api", type=int, default=0, help="0: universe 엑셀 사용(권장), 1: 랭킹 API 사용")
    ap.add_argument("--universe-excel", default="turnover_universe.xlsx", help="turnover_universe.xlsx 경로(use-api=0일 때)")
    ap.add_argument("--sleep", type=float, default=0.25, help="API 호출 간 대기(초)")
    ap.add_argument("--appkey", default=None, help="키움 APPKEY (인자>환경변수>하드코딩)")
    ap.add_argument("--secret", default=None, help="키움 SECRET (인자>환경변수>하드코딩)")
    args = ap.parse_args()

    # 기간 결정
    if args.from_date and args.to_date:
        start, end = args.from_date, args.to_date
    else:
        start, end = default_window_1y(args.to_date)

    # 키 우선순위: 인자 > 환경변수 > 하드코딩
    appkey = args.appkey or os.getenv("KIWOOM_APPKEY") or APPKEY_DEFAULT
    secret = args.secret or os.getenv("KIWOOM_SECRET") or SECRET_DEFAULT

    sess = KiwoomSession(appkey, secret)

    # SUMMARY → DAILY 순서로 갱신
    df_sum = update_summary(sess, args.out, start, end, args.threshold_eok, args.use_api, args.universe_excel, args.sleep)
    update_daily(sess, args.out, df_sum, start, end, args.sleep)
    print(f"[DONE] window={start}~{end} → {args.out}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Fatal: {e}")
        sys.exit(1)
