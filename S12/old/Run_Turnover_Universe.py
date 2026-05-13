#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Run_Turnover_Universe — FINAL v2
- 정확한 '해당 날짜' 차트 매칭
- 재구성(reconstruct) 시 실행 구간 truncate 후 전일자 재계산
- 최근 N 영업일(--bdays) + 출력 파일(--out)
- (날짜, 티커) 중복 제거 유지
"""

from __future__ import annotations
import os, sys, argparse, logging, datetime as dt, time
from typing import List, Dict, Any, Tuple, Optional

import requests
import pandas as pd
import numpy as np

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# --------------------------- 설정 ---------------------------
SHEET_UNIVERSE = "universe"
EXCEL_PATH     = os.path.abspath("./turnover_universe.xlsx")

# 요청: 코드에 큰따옴표 포함된 기본키(보안 무시)
APPKEY_DEFAULT = "\"IweTdkYa8JWDUOa8NohVSVeOiJ1THDGd_2x050A8XcU\""
SECRET_DEFAULT = "\"eazu-jPNJpAsIVkaUTh3_88gUvXrCMJCwGF2AYRtBJs\""

KIWOOM_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
KIWOOM_BASE_URL  = "https://api.kiwoom.com"
ENDPOINT_RANK    = "/api/dostk/rkinfo"   # ka10032
ENDPOINT_CHART   = "/api/dostk/chart"    # ka10081
API_ID_RANK      = "ka10032"
API_ID_CHART     = "ka10081"

RANK_FIXED_BODY = {
    "qry_tp": "2",            # 전일 거래대금 상위
    "rank_strt": "0",
    "rank_end":  "100",
    "stex_tp":   "3",         # 통합
    "mang_stk_incls": "1",
}
MARKETS = ["001","101"]       # KRX, KOSDAQ

THRESHOLD_EOK_DEFAULT = 5000.0
EXCLUDE_KEYWORDS = [
    "KODEX","TIGER","KBSTAR","KOSEF","ARIRANG","HANARO","SOL","TREX","ACE",
    "인버스","레버리지","선물","ETF","ETN","지수"
]

# --------------------------- 로깅 ---------------------------
logger = logging.getLogger("turnover_universe")
_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S"))
logger.addHandler(_handler)
logger.setLevel(logging.INFO)

# --------------------------- 유틸/엑셀 ---------------------------
def normalize_ticker(t: Any) -> str:
    if t is None: return ""
    s = str(t).strip()
    if "_" in s: s = s.split("_")[0]
    s = "".join(ch for ch in s if ch.isdigit())
    return s.zfill(6)[:6] if s else ""

def ensure_book(path: str) -> None:
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_UNIVERSE
        ws.append(["날짜","티커","종목명","거래대금(억)"])
        wb.save(path)

def auto_fit_borders(path: str, sheet: str) -> None:
    try:
        wb = load_workbook(path)
        ws = wb[sheet]
        for ci, col in enumerate(ws.columns, start=1):
            width = max(len("" if c.value is None else str(c.value)) for c in col) + 8
            ws.column_dimensions[get_column_letter(ci)].width = width
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value not in (None,""):
                    cell.border = border
        wb.save(path)
    except Exception:
        pass

def read_universe(path: str) -> pd.DataFrame:
    ensure_book(path)
    try:
        df = pd.read_excel(path, sheet_name=SHEET_UNIVERSE, dtype={"티커": str})
    except Exception:
        df = pd.DataFrame(columns=["날짜","티커","종목명","거래대금(억)"])
    if not df.empty:
        df["티커"] = df["티커"].map(normalize_ticker)
        df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.date
        df = df.dropna(subset=["날짜"])
    return df

def save_universe(path: str, df: pd.DataFrame) -> None:
    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            df.to_excel(w, index=False, sheet_name=SHEET_UNIVERSE)
    except Exception:
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name=SHEET_UNIVERSE)
    auto_fit_borders(path, SHEET_UNIVERSE)

def append_universe_rows(path: str, rows: List[Tuple[str,str,str,float]]) -> None:
    df_old = read_universe(path)
    df_new = pd.DataFrame(rows, columns=["날짜","티커","종목명","거래대금(억)"])
    if df_new.empty:
        return
    df_new["티커"] = df_new["티커"].map(normalize_ticker)
    df_new["날짜"] = pd.to_datetime(df_new["날짜"], errors="coerce").dt.date
    df_all = pd.concat([df_old, df_new], ignore_index=True)
    # (날짜, 티커) 기준 중복 제거
    df_all = df_all.drop_duplicates(subset=["날짜","티커"], keep="last")
    save_universe(path, df_all)

def truncate_date_range(path: str, start: str, end: str) -> None:
    df = read_universe(path)
    if df.empty: 
        return
    s = pd.to_datetime(start).date()
    e = pd.to_datetime(end).date()
    keep = ~((df["날짜"] >= s) & (df["날짜"] <= e))
    save_universe(path, df.loc[keep].copy())

# --------------------------- REST ---------------------------
def acquire_token(appkey: str, secret: str) -> str:
    headers = {"Content-Type":"application/json;charset=UTF-8"}
    body = {"grant_type":"client_credentials", "appkey": appkey.strip('"'), "secretkey": secret.strip('"')}
    r = requests.post(KIWOOM_TOKEN_URL, headers=headers, json=body, timeout=20)
    if r.status_code != 200:
        logger.warning("[AUTH] status=%s body=%s", r.status_code, r.text)
    r.raise_for_status()
    data = r.json()
    token = data.get("token") or data.get("access_token")
    if not token:
        raise RuntimeError("token missing")
    logger.info("[AUTH] token acquired")
    return token

def _post_with_retry(url: str, headers: dict, body: dict, max_retry: int = 6) -> dict:
    last = None
    for i in range(max_retry):
        try:
            r = requests.post(url, headers=headers, json=body, timeout=20)
            if r.status_code == 429:
                ra = r.headers.get("Retry-After")
                slp = float(ra) if (ra and str(ra).isdigit()) else (0.5 * (2 ** i))
                time.sleep(slp); continue
            if 500 <= r.status_code < 600:
                time.sleep(0.5 * (2 ** i)); continue
            r.raise_for_status()
            return r.json()
        except requests.RequestException as e:
            last = e
            time.sleep(0.5 * (2 ** i))
    if last: raise last
    raise RuntimeError("unknown error")

def post_rank(token: str, body: dict) -> dict:
    headers = {"authorization": f"Bearer {token}", "Content-Type":"application/json;charset=UTF-8",
               "api-id": API_ID_RANK, "cont-yn":"N", "next-key":""}
    return _post_with_retry(KIWOOM_BASE_URL + ENDPOINT_RANK, headers, body)

def post_chart(token: str, body: dict) -> dict:
    headers = {"authorization": f"Bearer {token}", "Content-Type":"application/json;charset=UTF-8",
               "api-id": API_ID_CHART, "cont-yn":"N", "next-key":""}
    return _post_with_retry(KIWOOM_BASE_URL + ENDPOINT_CHART, headers, body)

# --------------------------- 파싱/필터 ---------------------------
def is_excluded_name(name: str) -> bool:
    if not name: return True
    up = str(name).upper()
    return any(kw in up for kw in [k.upper() for k in EXCLUDE_KEYWORDS])

def _first_key(d: dict, cands: list, default=None):
    for k in cands:
        if k in d and d.get(k) not in (None, ""):
            return d.get(k)
    return default

def parse_rank_rows(resp: dict) -> List[Dict[str, Any]]:
    rows = []
    container = None
    for k, v in resp.items():
        if isinstance(v, list):
            container = v; break
    if not container: 
        return rows

    def to_float(x):
        try: return float(str(x).replace(",",""))
        except: return np.nan

    for rec in container:
        tkr = _first_key(rec, ["stk_cd","stck_shrn_iscd","ISU_SRT_CD","isu_cd"])
        name= _first_key(rec, ["stk_kor_isnm","stk_kor_nm","ISU_ABBRV","stk_nm","isu_nm"])
        amt = _first_key(rec, ["tvol_tamt","TOT_TR_PRC","tot_tr_prc","trdval","trd_prc","trde_prica","trdval_amt"])
        if not tkr or not name: 
            continue
        val = to_float(amt)
        if np.isnan(val): 
            continue
        rows.append({"ticker": normalize_ticker(tkr), "name": str(name).strip(), "turnover_100man": val})
    return rows

def merge_markets_unified(rows_by_mkt: List[List[Dict[str, Any]]]) -> pd.DataFrame:
    flat = [r for part in rows_by_mkt for r in part]
    if not flat:
        return pd.DataFrame(columns=["ticker","name","turnover_100man","turnover_eok"])
    df = pd.DataFrame(flat)
    df = df.groupby(["ticker","name"], as_index=False)["turnover_100man"].sum()
    # 단위 고정: ×0.01 → 억
    df["turnover_eok"] = df["turnover_100man"] * 0.01
    return df

def filter_top(df: pd.DataFrame, threshold_eok: float) -> pd.DataFrame:
    if df.empty: return df
    mask = ~df["name"].apply(is_excluded_name)
    out = df.loc[mask].copy()
    out = out[out["turnover_eok"] >= threshold_eok].copy()
    return out.sort_values("turnover_eok", ascending=False).reset_index(drop=True)

def print_head(df: pd.DataFrame, ymd: str, head_n: int = 10) -> None:
    if df.empty:
        logger.info("[TOP %s] (empty)", ymd); return
    head = df.head(head_n).copy()
    head["fmt"] = head["turnover_eok"].map(lambda x: f"{int(round(x)):,}")
    logger.info("[TOP %s] 5,000억 이상 (head %d)", ymd, head_n)
    for _, r in head.iterrows():
        logger.info("%s  %s  %s억", r["ticker"], r["name"], r["fmt"])

# --------------------------- 차트: 정확 날짜 매칭 ---------------------------
def _norm_date8(x: Any) -> str:
    if x in (None, ""): return ""
    s = str(x).replace("-", "").replace(".", "").strip()
    if len(s) < 8: return ""
    return f"{s[:4]}-{s[4:6]}-{s[6:8]}"

def day_turnover_eok_from_chart(token: str, ticker: str, ymd: str) -> float:
    # 같은 날만 요청하지만, 일부 벤더는 인접일도 같이 줄 수 있으니 날짜 매칭 로직으로 필터
    body = {"stk_cd": ticker, "base_dt": ymd.replace("-", ""), "end_dt": ymd.replace("-", ""), "upd_stkpc_tp":"1"}
    try:
        resp = post_chart(token, body)
    except Exception:
        return 0.0

    container = None
    for k, v in resp.items():
        if isinstance(v, list):
            container = v; break
    if not container: 
        return 0.0

    rec_y = None
    for rec in container:
        d_raw = _first_key(rec, ["dt","stck_bsop_date","TRD_DD","base_dt","biz_dt"])
        if _norm_date8(d_raw) == ymd:
            rec_y = rec; break
    if rec_y is None:
        return 0.0

    def f(x):
        try: return float(str(x).replace(",",""))
        except: return np.nan

    close = f(_first_key(rec_y, ["cur_prc","stck_clpr","close","END_PRC","CLOSE_PRC"]))
    vol   = f(_first_key(rec_y, ["trd_qty","stck_tvol","volume","ACC_TRDVOL"]))
    trdv  = f(_first_key(rec_y, ["trdval","trdval_amt","acc_trdval","ACC_TRDVAL","tot_tr_prc","TOT_TR_PRC"]))

    if not np.isnan(trdv) and trdv > 0:
        return trdv / 100_000_000.0
    if not (np.isnan(close) or np.isnan(vol)):
        return (close * vol) / 100_000_000.0
    return 0.0

# --------------------------- 날짜 ---------------------------
def business_days(start: dt.date, end: dt.date) -> List[str]:
    return [d.strftime("%Y-%m-%d") for d in pd.bdate_range(start=start, end=end)]

# --------------------------- 메인 루프 ---------------------------
def run_range(token: str, start: str, end: str, threshold_eok: float, reconstruct: int, seed_top: int) -> None:
    s = dt.datetime.strptime(start, "%Y-%m-%d").date()
    e = dt.datetime.strptime(end, "%Y-%m-%d").date()
    days = business_days(s, e)
    logger.info("[RANGE] %s ~ %s 영업일 %d일", start, end, len(days))

    total_saved = 0
    stale_cnt = 0
    last_sig = None

    seed_tickers: List[str] = []
    seed_name_map: Dict[str, str] = {}

    for i, ymd in enumerate(days, start=1):
        logger.info("[%d/%d] %s 처리 중...", i, len(days), ymd)

        # 1) 랭킹 시도
        rows_by_mkt: List[List[Dict[str, Any]]] = []
        for m in MARKETS:
            body = dict(RANK_FIXED_BODY)
            body["mrkt_tp"] = m
            body["bas_dd"]  = ymd.replace("-", "")
            try:
                resp = post_rank(token, body)
                rows_by_mkt.append(parse_rank_rows(resp))
            except Exception as e_req:
                logger.warning("[RANK][%s] market=%s error: %s", ymd, m, e_req)
                rows_by_mkt.append([])
        df_rank = merge_markets_unified(rows_by_mkt)
        df_f = filter_top(df_rank, threshold_eok)

        # 첫날 seed 확보
        if i == 1 and not df_f.empty:
            head = df_f.head(seed_top).copy()
            seed_tickers = head["ticker"].tolist()
            seed_name_map = {r["ticker"]: r["name"] for _, r in head.iterrows()}

        # 반복 감지(상위10 티커+합계)
        sig = (tuple(df_f.head(10)["ticker"].tolist()), float(df_f["turnover_eok"].sum() or 0.0))
        stale_cnt = (stale_cnt + 1) if (last_sig == sig) else 0
        last_sig = sig

        if stale_cnt >= 2:
            if reconstruct:
                logger.warning("[RECONSTRUCT] 랭킹 과거일 미지원 감지 → %s~%s 구간 차트 재구성", start, end)
                # 0) 오염 행 제거
                truncate_date_range(EXCEL_PATH, start, end)

                # 1) seed 보강(혹시 비어있으면 최신일로 다시)
                if not seed_tickers:
                    latest_rows = []
                    for m in MARKETS:
                        body = dict(RANK_FIXED_BODY); body["mrkt_tp"]=m; body["bas_dd"]=end.replace("-", "")
                        try:
                            resp = post_rank(token, body)
                            latest_rows.append(parse_rank_rows(resp))
                        except Exception:
                            latest_rows.append([])
                    df_latest = filter_top(merge_markets_unified(latest_rows), threshold_eok)
                    head = df_latest.head(seed_top).copy()
                    seed_tickers = head["ticker"].tolist()
                    seed_name_map = {r["ticker"]: r["name"] for _, r in head.iterrows()}

                # 2) 전 구간 재계산
                for d in days:
                    rows = []
                    for t in seed_tickers:
                        eok = day_turnover_eok_from_chart(token, t, d)
                        if eok >= threshold_eok:
                            rows.append((d, t, seed_name_map.get(t,""), float(eok)))
                    if rows:
                        append_universe_rows(EXCEL_PATH, rows)
                        total_saved += len(rows)
                        logger.info("[SAVE][RECON] %s rows=%d → %s", d, len(rows), EXCEL_PATH)
                    else:
                        logger.info("[SAVE][RECON] %s 저장할 행 없음", d)
                logger.info("[DONE][RECON] 차트 기반 재구성 완료 (총 저장 %d행)", total_saved)
            else:
                logger.warning("[SKIP] 랭킹 응답이 과거일을 반영하지 않음 → 중단")
            break

        # 2) 랭킹이 정상적이면 저장
        print_head(df_f, ymd)
        rows = [(ymd, t, n, float(v)) for t, n, v in df_f[["ticker","name","turnover_eok"]].itertuples(index=False, name=None)]
        if rows:
            append_universe_rows(EXCEL_PATH, rows)
            total_saved += len(rows)
            logger.info("[SAVE] %s rows=%d → %s", ymd, len(rows), EXCEL_PATH)
        else:
            logger.info("[SAVE] %s 저장할 행 없음", ymd)

    logger.info("[DONE] 범위 처리 완료: 총 저장 %d행", total_saved)

# --------------------------- 엔트리 ---------------------------
def main():
    parser = argparse.ArgumentParser(description="Turnover Universe Builder — FINAL v2")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--bdays", type=int, default=None, help="오늘 기준 최근 N 영업일만 처리 (예: 14)")
    parser.add_argument("--days", type=int, default=None, help="오늘 기준 최근 N일(달력일)")
    parser.add_argument("--from", dest="from_date", default=None, help="YYYY-MM-DD")
    parser.add_argument("--to",   dest="to_date",   default=None, help="YYYY-MM-DD")
    parser.add_argument("--out",  default="turnover_universe.xlsx", help="출력 엑셀 경로")
    parser.add_argument("--threshold-eok", type=float, default=THRESHOLD_EOK_DEFAULT, help="거래대금 임계값(억)")
    parser.add_argument("--reconstruct", type=int, default=1, help="랭킹 반복 시 차트 재구성(1=on,0=off)")
    parser.add_argument("--seed-top", type=int, default=200, help="최신일 seed 티커 개수")
    parser.add_argument("--appkey", default=None)
    parser.add_argument("--secret", default=None)
    args = parser.parse_args()

    if args.verbose: logger.setLevel(logging.DEBUG)

    appkey = (args.appkey or os.getenv("KIWOOM_APPKEY") or APPKEY_DEFAULT)
    secret = (args.secret or os.getenv("KIWOOM_SECRET") or SECRET_DEFAULT)

    global EXCEL_PATH
    EXCEL_PATH = os.path.abspath(args.out)

    logger.info("Python %s", sys.version.replace("\n"," "))
    logger.info("[START] turnover scan")

    token = acquire_token(appkey, secret)

    # 범위 결정: bdays > from/to > days > default(30달력일)
    if args.bdays:
        e = dt.date.today()
        bdays = pd.bdate_range(end=e, periods=int(args.bdays))
        s = bdays[0].date()
        run_range(token, s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d"),
                  args.threshold_eok, args.reconstruct, args.seed_top)
    elif args.from_date and args.to_date:
        run_range(token, args.from_date, args.to_date, args.threshold_eok, args.reconstruct, args.seed_top)
    elif args.days:
        e = dt.date.today(); s = e - dt.timedelta(days=int(args.days))
        run_range(token, s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d"),
                  args.threshold_eok, args.reconstruct, args.seed_top)
    else:
        e = dt.date.today(); s = e - dt.timedelta(days=30)
        run_range(token, s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d"),
                  args.threshold_eok, args.reconstruct, args.seed_top)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.exception("Fatal error: %s", e)
        sys.exit(1)
