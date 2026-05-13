from __future__ import annotations
import os
import pathlib
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config.adapters import BinanceClient
from universe_selector import get_top30_coins, get_top30_symbols
from core.phase1_5_core import run_phase1_5_simulation


OUTPUT_DIR = pathlib.Path("debug")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def convert_csv_to_excel(csv_path: pathlib.Path) -> pathlib.Path:
    """CSV 파일을 Excel 파일로 변환하고 A열 너비 조정 및 1행 고정"""
    excel_path = csv_path.with_suffix('.xlsx')
    
    # CSV 파일 읽기
    df = pd.read_csv(csv_path)
    
    # Excel 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Debug Data"
    
    # DataFrame을 Excel에 추가
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # A열 너비 조정 (날짜 컬럼)
    ws.column_dimensions['A'].width = 15
    
    # 1행 고정 (헤더 고정)
    ws.freeze_panes = 'A2'
    
    # 헤더 스타일링
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Excel 파일 저장
    wb.save(excel_path)
    return excel_path


def build_all(limit_days: int = 1200, symbols: Optional[list[str]] = None, top_n: int = 100) -> list[str]:
    """
    Build per-symbol debug CSVs for Top N (or provided symbols).
    - Downloads 일봉 OHLCV from Binance.
    - Computes Phase 1.5 debug table (H 루프 보정/리셋 포함).
    - Saves to debug/{SYMBOL}_debug.csv
    - Excludes stablecoins, wrapped tokens, and unsupported symbols.
    Returns list of produced file paths (as str).
    """
    client = BinanceClient()
    
    if symbols:
        syms = symbols
        print(f"[INFO] Using provided symbols: {len(syms)}개")
    else:
        # Top N 코인 리스트 가져오기
        coins = get_top30_coins()  # universe_selector에서 TOP_N=100으로 설정됨
        syms = [coin["Symbol"] for coin in coins[:top_n]]
        print(f"[INFO] Using Top {top_n} coins: {len(syms)}개")
    
    # Binance API 미지원 심볼 제외 (스테이블코인, 래핑 토큰, 데이터 없음)
    exclude_symbols = {
        "USDTUSDT", "USDCUSDT", "USDEUSDT", "USDSUSDT", "DAIUSDT",  # 스테이블코인
        "WBTCUSDT", "WBETHUSDT", "WEETHUSDT", "STETHUSDT", "WSTETHUSDT",  # 래핑 토큰
        "FIGR_HELOCUSDT", "HYPEUSDT", "LEOUSDT", "USDT0USDT", "SUSDEUSDT",  # API 에러
        "MUSDT", "OKBUSDT", "WLFIUSDT", "BGBUSDT", "MNTUSDT", "CROUSDT",  # 데이터 없음
        "USD1USDT", "BFUSDUSDT", "BNSOLUSDT", "BNBUSDT", "ENAUSDT"  # 추가 제외 코인
    }
    syms = [s for s in syms if s not in exclude_symbols]

    produced: list[str] = []
    total_syms = len(syms)
    successful = 0
    failed = 0
    
    print(f"\n{'='*60}")
    print(f"Phase 1.5 Debug Builder - Top {top_n} 코인")
    print(f"{'='*60}")
    print(f"총 대상: {total_syms}개 코인")
    print(f"데이터 기간: {limit_days}일")
    print(f"{'='*60}\n")
    
    for i, sym in enumerate(syms, 1):
        sym_name = sym.replace("USDT", "")
        print(f"[{i:3d}/{total_syms}] {sym_name:<8} ({sym}) 처리 중...", end=" ")
        
        try:
            # OHLC 데이터 가져오기
            df = client.get_ohlc_daily(sym, limit=limit_days)
            if df.empty:
                print("FAIL 데이터 없음")
                failed += 1
                continue
            
            # Convert DataFrame to list of dictionaries for run_phase1_5_simulation
            ohlc_data = []
            for _, row in df.iterrows():
                # Convert date to timestamp in milliseconds
                timestamp_ms = int(row['date'].timestamp() * 1000)
                
                # 2025-10-09 날짜의 저가 데이터를 종가로 대체 (이상 데이터 보정)
                date_str = row['date'].strftime('%Y-%m-%d')
                low_value = float(row['close']) if date_str == '2025-10-09' else float(row['low'])
                
                ohlc_data.append({
                    'closeTime': timestamp_ms,  # Use closeTime as expected by run_phase1_5_simulation
                    'open': float(row['open']),
                    'high': float(row['high']),
                    'low': low_value,
                    'close': float(row['close']),
                    'volume': float(row['volume'])
                })
            
            # Phase 1.5 시뮬레이션 실행
            out_path = OUTPUT_DIR / f"{sym_name}_debug.csv"
            
            run_phase1_5_simulation(
                symbol=sym,
                ohlc=ohlc_data,
                seed_H=None,  # H는 첫 사이클 시작 시 자동 설정
                out_csv=out_path,
                limit_days=limit_days
            )
            
            # CSV를 Excel로 변환
            excel_path = convert_csv_to_excel(out_path)
            
            produced.append(str(excel_path))
            successful += 1
            print(f"OK 완료 ({len(ohlc_data)}일 데이터, Excel 변환)")
            
        except Exception as e:
            failed += 1
            print(f"FAIL 실패: {str(e)[:50]}...")
            continue
    
    # 결과 요약
    print(f"\n{'='*60}")
    print(f"처리 완료!")
    print(f"OK 성공: {successful}개")
    print(f"FAIL 실패: {failed}개")
    print(f"저장 위치: {OUTPUT_DIR.resolve()}")
    print(f"{'='*60}")
    
    return produced


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Phase 1.5 Debug Builder - Top N 코인 일괄 처리")
    parser.add_argument("--top-n", type=int, default=100, help="처리할 Top N 코인 수 (기본: 100)")
    parser.add_argument("--limit-days", type=int, default=1200, help="데이터 기간 (기본: 1200일)")
    parser.add_argument("--symbols", nargs="+", help="특정 심볼들만 처리 (예: BTCUSDT ETHUSDT)")
    
    args = parser.parse_args()
    
    if args.symbols:
        files = build_all(limit_days=args.limit_days, symbols=args.symbols)
    else:
        files = build_all(limit_days=args.limit_days, top_n=args.top_n)
    
    print(f"\n완료! 총 {len(files)}개 파일 생성 완료!")
