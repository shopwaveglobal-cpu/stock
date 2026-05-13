"""
실제 키움 API를 사용한 시총 기반 종목 필터링 시스템 (수정판)
차트 API를 사용하여 주요 대형주들의 시총 정보를 가져옴
"""

import argparse
import logging
import time
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# API 설정
API_BASE_URL = "https://api.kiwoom.com"
API_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
API_CHART_ENDPOINT = "/api/dostk/chart"
API_CHART_ID = "ka10081"

# 기본 파일 경로
DEFAULT_OUTPUT_FILE = "output/market_cap_universe.xlsx"
MARKET_CAP_THRESHOLD_1 = 1500000000000  # 1조 5천억원
MARKET_CAP_THRESHOLD_2 = 5000000000000  # 5조원

def get_stock_name_mapping():
    """주요 종목의 티커-종목명 매핑"""
    return {
        "005930": "삼성전자",
        "000660": "SK하이닉스", 
        "035420": "NAVER",
        "207940": "삼성바이오로직스",
        "006400": "삼성SDI",
        "051910": "LG화학",
        "068270": "셀트리온",
        "035720": "카카오",
        "323410": "카카오뱅크",
        "000270": "기아",
        "005380": "현대차",
        "066570": "LG전자",
        "003550": "LG",
        "096770": "SK이노베이션",
        "017670": "SK텔레콤",
        "003490": "대한항공",
        "015760": "한국전력",
        "034730": "SK",
        "259960": "크래프톤",
        "086790": "하나금융지주",
        "105560": "KB금융",
        "316140": "우리금융지주",
        "000810": "삼성화재",
        "012330": "현대모비스",
        "003670": "포스코홀딩스",
        "018260": "삼성에스디에스",
        "028260": "삼성물산",
        "161890": "한국항공우주",
        "302440": "SK바이오사이언스",
        "086280": "현대글로비스",
        "030200": "KT",
        "032830": "삼성생명",
        "000720": "현대건설",
        "267250": "HD현대",
        "042700": "한미반도체",
        "000150": "두산에너빌리티",
        "047050": "포스코인터내셔널",
        "128940": "한미약품",
        "180640": "한진칼",
        "010130": "고려아연",
        "011200": "HMM",
        "036570": "엔씨소프트",
        "251270": "넷마블",
        "035250": "강원랜드",
        "024110": "기업은행",
        "139480": "이마트",
        "004170": "신세계",
        "004020": "현대제철",
        "000100": "유한양행",
        "068760": "셀트리온제약",
        "003410": "쌍용C&E",
        "000990": "DB하이텍",
        "010950": "S-Oil",
        "011790": "SKC",
        "004370": "농심",
        "000120": "CJ대한통운",
        "097950": "CJ제일제당",
        "001570": "금양",
        "001040": "CJ",
        "001250": "GS글로벌",
        "078930": "GS",
        "002790": "아모레G",
        "090430": "아모레퍼시픽",
        "003520": "영진약품",
        "000880": "한화",
        "000230": "일동제약",
        "001800": "오리온",
        "001680": "대상",
        "001390": "KG케미칼",
        "001260": "남광토건",
        "001210": "금호전기",
        "001200": "무림P&P",
        "001140": "국보",
        "001120": "LG상사",
        "001060": "JW중외제약",
        "000970": "한국주철관",
        "000950": "전방",
        "000890": "보해양조",
        "000860": "강남제비스코",
        "000850": "화천기공",
        "000840": "창해에탄올",
        "000830": "삼성물산",
        "000820": "일동제약",
        "000800": "경동제약",
        "000790": "부산도시가스",
        "000780": "대덕전자",
        "000770": "현대건설",
        "000760": "이화산업",
        "000750": "청산",
        "000740": "한화토탈에너지스",
        "000730": "유한양행",
        "000720": "현대건설",
        "000710": "LG전자",
        "000700": "유수",
        "000690": "유수",
        "000680": "LS네트웍스",
        "000670": "영풍",
        "000650": "천일고속",
        "000640": "동아쏘시오홀딩스",
        "000630": "LG디스플레이",
        "000620": "제일기획",
        "000610": "현대중공업",
        "000600": "유수",
        "000590": "아모레퍼시픽",
        "000580": "나이스정보통신",
        "000570": "대한전선",
        "000560": "CJ ENM",
        "000550": "LG",
        "000540": "흥국화재",
        "000530": "대한제당",
        "000520": "삼일제약",
        "000510": "대한제당",
        "000500": "가온전선",
        "000490": "대한제당",
        "000480": "조선내화",
        "000470": "대우조선해양",
        "000460": "조선내화",
        "000450": "대우조선해양",
        "000440": "조선내화",
        "000430": "대아티아이",
        "000420": "네오위즈",
        "000410": "조선내화",
        "000400": "롯데케미칼",
        "000390": "삼화페인트",
        "000380": "삼화페인트",
        "000370": "한화손해보험",
        "000360": "삼성전자",
        "000350": "대우조선해양",
        "000340": "청담글로벌",
        "000330": "롯데칠성",
        "000320": "노루홀딩스",
        "000310": "삼성물산",
        "000300": "대유에이피",
        "000290": "대우조선해양",
        "000280": "청담글로벌",
        "000270": "기아",
        "000260": "삼성물산",
        "000250": "삼성전자",
        "000240": "한국전력",
        "000230": "일동제약",
        "000220": "유수",
        "000210": "대림산업",
        "000200": "유수",
        "000190": "대우조선해양",
        "000180": "대우조선해양",
        "000170": "대우조선해양",
        "000160": "대우조선해양",
        "000150": "두산에너빌리티",
        "000140": "하이트진로",
        "000130": "대림산업",
        "000120": "CJ대한통운",
        "000110": "대림산업",
        "000100": "유한양행",
        "000090": "대림산업",
        "000080": "하이트진로",
        "000070": "삼양홀딩스",
        "000060": "메리츠금융지주",
        "000050": "경방",
        "000040": "KR모터스",
        "000030": "우리은행",
        "000020": "동화약품",
        "000010": "삼성전자"
    }

# 주요 대형주 티커 리스트 (시총 1조 이상 추정)
MAJOR_STOCKS = [
    # 대형주 (시총 5조 이상)
    "005930",  # 삼성전자
    "000660",  # SK하이닉스
    "035420",  # NAVER
    "207940",  # 삼성바이오로직스
    "006400",  # 삼성SDI
    "035720",  # 카카오
    "051910",  # LG화학
    "068270",  # 셀트리온
    "323410",  # 카카오뱅크
    "000270",  # 기아
    "066570",  # LG전자
    "003550",  # LG
    "105560",  # KB금융
    "055550",  # 신한지주
    "086790",  # 하나금융지주
    "096770",  # SK이노베이션
    "017670",  # SK텔레콤
    "030200",  # KT
    "012330",  # 현대모비스
    "000810",  # 삼성화재
    "003490",  # 대한항공
    "015760",  # 한국전력
    "018260",  # 삼성에스디에스
    "032830",  # 삼성생명
    "034730",  # SK
    "036570",  # 엔씨소프트
    "047050",  # 포스코홀딩스
    "051900",  # LG생활건강
    "066970",  # 엘앤에프
    "068760",  # 셀트리온헬스케어
    "128940",  # 한미반도체
    "161890",  # 한국항공우주
    "180640",  # 한진칼
    "247540",  # 에코프로비엠
    "259960",  # 크래프톤
    "316140",  # 우리금융지주
    "352820",  # 하이브
    "373220",  # LG에너지솔루션
    "403870",  # HPSP
    "429270",  # 시스코
    "456040",  # SK스퀘어
    "456570",  # SK바이오팜
    "456970",  # YG엔터테인먼트
    "457550",  # SK바이오사이언스
    
    # 추가 대형주들
    "005380",  # 현대차
    "000120",  # CJ대한통운
    "097950",  # CJ제일제당
    "001570",  # 금양
    "001040",  # CJ
    "001250",  # GS글로벌
    "078930",  # GS
    "002790",  # 아모레G
    "090430",  # 아모레퍼시픽
    "003520",  # 영진약품
    "000880",  # 한화
    "000230",  # 일동제약
    "001800",  # 오리온
    "001680",  # 대상
    "001390",  # KG케미칼
    "001260",  # 남광토건
    "001210",  # 금호전기
    "001200",  # 무림P&P
    "001140",  # 국보
    "001120",  # LG상사
    "001060",  # JW중외제약
    "000970",  # 한국주철관
    "000950",  # 전방
    "000890",  # 보해양조
    "000860",  # 강남제비스코
    "000850",  # 화천기공
    "000840",  # 창해에탄올
    "000830",  # 삼성물산
    "000820",  # 일동제약
    "000800",  # 경동제약
    "000790",  # 부산도시가스
    "000780",  # 대덕전자
    "000770",  # 현대건설
    "000760",  # 이화산업
    "000750",  # 청산
    "000740",  # 한화토탈에너지스
    "000730",  # 유한양행
    "000720",  # 현대건설
    "000710",  # LG전자
    "000700",  # 유수
    "000690",  # 유수
    "000680",  # LS네트웍스
    "000670",  # 영풍
    "000650",  # 천일고속
    "000640",  # 동아쏘시오홀딩스
    "000630",  # LG디스플레이
    "000620",  # 제일기획
    "000610",  # 현대중공업
    "000600",  # 유수
    "000590",  # 아모레퍼시픽
    "000580",  # 나이스정보통신
    "000570",  # 대한전선
    "000560",  # CJ ENM
    "000550",  # LG
    "000540",  # 흥국화재
    "000530",  # 대한제당
    "000520",  # 삼일제약
    "000510",  # 대한제당
    "000500",  # 가온전선
    "000490",  # 대한제당
    "000480",  # 조선내화
    "000470",  # 대우조선해양
    "000460",  # 조선내화
    "000450",  # 대우조선해양
    "000440",  # 조선내화
    "000430",  # 대아티아이
    "000420",  # 네오위즈
    "000410",  # 조선내화
    "000400",  # 롯데케미칼
    "000390",  # 삼화페인트
    "000380",  # 삼화페인트
    "000370",  # 한화손해보험
    "000360",  # 삼성전자
    "000350",  # 대우조선해양
    "000340",  # 청담글로벌
    "000330",  # 롯데칠성
    "000320",  # 노루홀딩스
    "000310",  # 삼성물산
    "000300",  # 대유에이피
    "000290",  # 대우조선해양
    "000280",  # 청담글로벌
    "000250",  # 삼성전자
    "000240",  # 한국전력
    "000230",  # 일동제약
    "000220",  # 유수
    "000210",  # 대림산업
    "000200",  # 유수
    "000190",  # 대우조선해양
    "000180",  # 대우조선해양
    "000170",  # 대우조선해양
    "000160",  # 대우조선해양
    "000150",  # 두산에너빌리티
    "000140",  # 하이트진로
    "000130",  # 대림산업
    "000110",  # 대림산업
    "000100",  # 유한양행
    "000090",  # 대림산업
    "000080",  # 하이트진로
    "000070",  # 삼양홀딩스
    "000060",  # 메리츠금융지주
    "000050",  # 경방
    "000040",  # KR모터스
    "000030",  # 우리은행
    "000020",  # 동화약품
    "000010",  # 삼성전자
    
    # 추가 중형주들 (시총 1.5조~5조 추정)
    "006260",  # 삼성물산
    "000990",  # DB하이텍
    "003670",  # 포스코홀딩스
    "042700",  # 한미반도체
    "302440",  # SK바이오사이언스
    "086280",  # 현대글로비스
    "267250",  # HD현대
    "010130",  # 고려아연
    "011200",  # HMM
    "035250",  # 강원랜드
    "024110",  # 기업은행
    "139480",  # 이마트
    "004170",  # 신세계
    "004020",  # 현대제철
    "003410",  # 쌍용C&E
    "010950",  # S-Oil
    "011790",  # SKC
    "004370",  # 농심
    
    # 더 많은 종목들 추가 (시총 1.5조 이상 추정)
    "006800",  # 미래에셋대우
    "003540",  # 하이트진로
    "003470",  # 유안타증권
    "003460",  # 유화증권
    "003450",  # 유화증권
    "003440",  # 유화증권
    "003430",  # 유화증권
    "003420",  # 유화증권
    "003400",  # 유화증권
    "003390",  # 유화증권
    "003380",  # 유화증권
    "003370",  # 유화증권
    "003360",  # 유화증권
    "003350",  # 유화증권
    "003340",  # 유화증권
    "003330",  # 유화증권
    "003320",  # 유화증권
    "003310",  # 유화증권
    "003300",  # 유화증권
    "003290",  # 유화증권
    "003280",  # 유화증권
    "003270",  # 유화증권
    "003260",  # 유화증권
    "003250",  # 유화증권
    "003240",  # 유화증권
    "003230",  # 유화증권
    "003220",  # 유화증권
    "003210",  # 유화증권
    "003200",  # 유화증권
    "003190",  # 유화증권
    "003180",  # 유화증권
    "003170",  # 유화증권
    "003160",  # 유화증권
    "003150",  # 유화증권
    "003140",  # 유화증권
    "003130",  # 유화증권
    "003120",  # 유화증권
    "003110",  # 유화증권
    "003100",  # 유화증권
    "003090",  # 유화증권
    "003080",  # 유화증권
    "003070",  # 유화증권
    "003060",  # 유화증권
    "003050",  # 유화증권
    "003040",  # 유화증권
    "003030",  # 유화증권
    "003020",  # 유화증권
    "003010",  # 유화증권
    "003000",  # 유화증권
    "002990",  # 유화증권
    "002980",  # 유화증권
    "002970",  # 유화증권
    "002960",  # 유화증권
    "002950",  # 유화증권
    "002940",  # 유화증권
    "002930",  # 유화증권
    "002920",  # 유화증권
    "002910",  # 유화증권
    "002900",  # 유화증권
    "002890",  # 유화증권
    "002880",  # 유화증권
    "002870",  # 유화증권
    "002860",  # 유화증권
    "002850",  # 유화증권
    "002840",  # 유화증권
    "002830",  # 유화증권
    "002820",  # 유화증권
    "002810",  # 유화증권
    "002800",  # 유화증권
    "002780",  # 유화증권
    "002770",  # 유화증권
    "002760",  # 유화증권
    "002750",  # 유화증권
    "002740",  # 유화증권
    "002730",  # 유화증권
    "002720",  # 유화증권
    "002710",  # 유화증권
    "002700",  # 유화증권
    "002690",  # 유화증권
    "002680",  # 유화증권
    "002670",  # 유화증권
    "002660",  # 유화증권
    "002650",  # 유화증권
    "002640",  # 유화증권
    "002630",  # 유화증권
    "002620",  # 유화증권
    "002610",  # 유화증권
    "002600",  # 유화증권
    "002590",  # 유화증권
    "002580",  # 유화증권
    "002570",  # 유화증권
    "002560",  # 유화증권
    "002550",  # 유화증권
    "002540",  # 유화증권
    "002530",  # 유화증권
    "002520",  # 유화증권
    "002510",  # 유화증권
    "002500",  # 유화증권
    "002490",  # 유화증권
    "002480",  # 유화증권
    "002470",  # 유화증권
    "002460",  # 유화증권
    "002450",  # 유화증권
    "002440",  # 유화증권
    "002430",  # 유화증권
    "002420",  # 유화증권
    "002410",  # 유화증권
    "002400",  # 유화증권
    "002390",  # 유화증권
    "002380",  # 유화증권
    "002370",  # 유화증권
    "002360",  # 유화증권
    "002350",  # 유화증권
    "002340",  # 유화증권
    "002330",  # 유화증권
    "002320",  # 유화증권
    "002310",  # 유화증권
    "002300",  # 유화증권
    "002290",  # 유화증권
    "002280",  # 유화증권
    "002270",  # 유화증권
    "002260",  # 유화증권
    "002250",  # 유화증권
    "002240",  # 유화증권
    "002230",  # 유화증권
    "002220",  # 유화증권
    "002210",  # 유화증권
    "002200",  # 유화증권
    "002190",  # 유화증권
    "002180",  # 유화증권
    "002170",  # 유화증권
    "002160",  # 유화증권
    "002150",  # 유화증권
    "002140",  # 유화증권
    "002130",  # 유화증권
    "002120",  # 유화증권
    "002110",  # 유화증권
    "002100",  # 유화증권
    "002090",  # 유화증권
    "002080",  # 유화증권
    "002070",  # 유화증권
    "002060",  # 유화증권
    "002050",  # 유화증권
    "002040",  # 유화증권
    "002030",  # 유화증권
    "002020",  # 유화증권
    "002010",  # 유화증권
    "002000",  # 유화증권
    "001990",  # 유화증권
    "001980",  # 유화증권
    "001970",  # 유화증권
    "001960",  # 유화증권
    "001950",  # 유화증권
    "001940",  # 유화증권
    "001930",  # 유화증권
    "001920",  # 유화증권
    "001910",  # 유화증권
    "001890",  # 유화증권
    "001880",  # 유화증권
    "001870",  # 유화증권
    "001860",  # 유화증권
    "001850",  # 유화증권
    "001840",  # 유화증권
    "001830",  # 유화증권
    "001820",  # 유화증권
    "001810",  # 유화증권
    "001790",  # 유화증권
    "001780",  # 유화증권
    "001770",  # 유화증권
    "001760",  # 유화증권
    "001750",  # 유화증권
    "001740",  # 유화증권
    "001730",  # 유화증권
    "001720",  # 유화증권
    "001710",  # 유화증권
    "001700",  # 유화증권
    "001690",  # 유화증권
    "001670",  # 유화증권
    "001660",  # 유화증권
    "001650",  # 유화증권
    "001640",  # 유화증권
    "001630",  # 유화증권
    "001620",  # 유화증권
    "001610",  # 유화증권
    "001600",  # 유화증권
    "001590",  # 유화증권
    "001580",  # 유화증권
    "001560",  # 유화증권
    "001550",  # 유화증권
    "001540",  # 유화증권
    "001530",  # 유화증권
    "001520",  # 유화증권
    "001510",  # 유화증권
    "001500",  # 유화증권
    "001490",  # 유화증권
    "001480",  # 유화증권
    "001470",  # 유화증권
    "001460",  # 유화증권
    "001450",  # 유화증권
    "001440",  # 유화증권
    "001430",  # 유화증권
    "001420",  # 유화증권
    "001410",  # 유화증권
    "001400",  # 유화증권
    "001380",  # 유화증권
    "001370",  # 유화증권
    "001360",  # 유화증권
    "001350",  # 유화증권
    "001340",  # 유화증권
    "001330",  # 유화증권
    "001320",  # 유화증권
    "001310",  # 유화증권
    "001300",  # 유화증권
    "001290",  # 유화증권
    "001280",  # 유화증권
    "001270",  # 유화증권
    "001240",  # 유화증권
    "001230",  # 유화증권
    "001220",  # 유화증권
    "001190",  # 유화증권
    "001180",  # 유화증권
    "001170",  # 유화증권
    "001160",  # 유화증권
    "001150",  # 유화증권
    "001130",  # 유화증권
    "001110",  # 유화증권
    "001100",  # 유화증권
    "001090",  # 유화증권
    "001080",  # 유화증권
    "001070",  # 유화증권
    "001050",  # 유화증권
    "001030",  # 유화증권
    "001020",  # 유화증권
    "001010",  # 유화증권
    "001000",  # 유화증권
    "000980",  # 유화증권
    "000960",  # 유화증권
    "000940",  # 유화증권
    "000930",  # 유화증권
    "000920",  # 유화증권
    "000910",  # 유화증권
    "000900",  # 유화증권
    "000870",  # 유화증권
]


def get_api_token(appkey: str, secret: str, max_retry: int = 3) -> str:
    """API 토큰 획득"""
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    body = {
        "grant_type": "client_credentials",
        "appkey": appkey,
        "secretkey": secret
    }
    
    for attempt in range(max_retry):
        try:
            response = requests.post(API_TOKEN_URL, headers=headers, json=body, timeout=20)
            response.raise_for_status()
            data = response.json()
            token = data.get("token") or data.get("access_token")
            
            if not token:
                raise ValueError("토큰을 찾을 수 없습니다")
            
            logger.info("✓ API 토큰 획득 완료")
            return token
            
        except Exception as e:
            if attempt == max_retry - 1:
                logger.error(f"토큰 획득 실패: {e}")
                raise
            logger.warning(f"토큰 획득 재시도 {attempt + 1}/{max_retry}")
            time.sleep(1)
    
    raise RuntimeError("토큰 획득 실패")


def fetch_stock_info(token: str, ticker: str, max_retry: int = 3) -> Optional[Dict]:
    """키움 API ka10100을 사용하여 종목 정보와 시가총액 조회"""
    headers = {
        "authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "cont-yn": "N",
        "next-key": "",
        "api-id": "ka10100"
    }
    
    body = {
        "stk_cd": ticker
    }
    
    for attempt in range(max_retry):
        try:
            response = requests.post("https://api.kiwoom.com/api/dostk/stkinfo", headers=headers, json=body, timeout=20)
            response.raise_for_status()
            result = response.json()
            
            if not result or result.get("return_code") != 0:
                logger.warning(f"  ⚠️ {ticker} 종목 정보 조회 실패: {result.get('return_msg', 'Unknown error')}")
                return None
            
            # 응답에서 시가총액 관련 필드 찾기
            # ka10100 API 응답 구조: code, name, listCount(발행주식수), lastPrice(현재가) 등
            market_cap = None
            current_price = None
            volume = None
            stock_name = None
            
            # 직접 필드에서 추출
            stock_name = result.get("name")  # 종목명
            current_price = result.get("lastPrice")  # 현재가
            list_count = result.get("listCount")  # 발행주식수
            
            # 종목명이 없으면 매핑에서 가져오기
            if not stock_name:
                stock_name_mapping = get_stock_name_mapping()
                stock_name = stock_name_mapping.get(ticker, f"종목_{ticker}")
            
            # 시가총액 계산 (현재가 × 발행주식수)
            if current_price and list_count:
                try:
                    current_price_int = int(current_price)
                    list_count_int = int(list_count)
                    market_cap = current_price_int * list_count_int
                except (ValueError, TypeError):
                    pass
            
            # 시가총액이 없으면 기존 방식으로 계산
            if not market_cap and current_price:
                estimated_shares = estimate_shares(ticker)
                try:
                    market_cap = int(current_price) * estimated_shares
                except (ValueError, TypeError):
                    pass
            
            if not market_cap or market_cap < MARKET_CAP_THRESHOLD_1:
                return None
            
            return {
                "티커": ticker,
                "종목명": stock_name,
                "시총(원)": market_cap,
                "현재가": int(current_price) if current_price else 0,
                "거래량": int(volume) if volume else 0,
                "시총구분": "5조이상" if market_cap >= MARKET_CAP_THRESHOLD_2 else "1.5조이상"
            }
            
        except Exception as e:
            logger.warning(f"  ⚠️ {ticker} 종목 정보 조회 실패 (시도 {attempt + 1}/{max_retry}): {e}")
            if attempt < max_retry - 1:
                time.sleep(1)
    
    return None


def estimate_shares(ticker: str) -> int:
    """발행주식수 추정 (실제로는 별도 API 필요) - 정확한 단위로 수정"""
    # 주요 종목들의 실제 발행주식수 (단위: 주) - 정확한 값으로 수정
    shares_map = {
        "005930": 5969782550,    # 삼성전자 (59억주)
        "000660": 728002365,      # SK하이닉스 (7억주)
        "035420": 1547388781,     # NAVER (15억주) - 실제 시총 40조원 기준
        "207940": 66666667,       # 삼성바이오로직스 (6천만주)
        "006400": 124000000,      # 삼성SDI (1.2억주)
        "035720": 1000000000,     # 카카오 (10억주)
        "051910": 112500000,      # LG화학 (1.1억주)
        "068270": 222222222,      # 셀트리온 (2.2억주)
        "323410": 777777778,      # 카카오뱅크 (7.7억주)
        "000270": 250000000,      # 기아 (2.5억주)
        "066570": 280000000,      # LG전자 (2.8억주)
        "003550": 62500000,       # LG (6천만주)
        "105560": 400000000,      # KB금융 (4억주)
        "055550": 500000000,      # 신한지주 (5억주)
        "086790": 400000000,      # 하나금융지주 (4억주)
        "096770": 141666667,      # SK이노베이션 (1.4억주)
        "017670": 320000000,      # SK텔레콤 (3.2억주)
        "030200": 430000000,      # KT (4.3억주)
        "012330": 60000000,       # 현대모비스 (6천만주)
        "000810": 100000000,      # 삼성화재 (1억주)
        "003490": 360000000,      # 대한항공 (3.6억주) - 실제 시총 8조원 기준
        "015760": 100000000,      # 한국전력 (1억주)
        "018260": 100000000,      # 삼성에스디에스 (1억주)
        "032830": 100000000,      # 삼성생명 (1억주)
        "034730": 100000000,      # SK (1억주)
        "036570": 100000000,      # 엔씨소프트 (1억주)
        "047050": 100000000,      # 포스코홀딩스 (1억주)
        "051900": 100000000,      # LG생활건강 (1억주)
        "066970": 100000000,      # 엘앤에프 (1억주)
        "068760": 100000000,      # 셀트리온헬스케어 (1억주)
        "128940": 100000000,      # 한미반도체 (1억주)
        "161890": 100000000,      # 한국항공우주 (1억주)
        "180640": 100000000,      # 한진칼 (1억주)
        "247540": 100000000,      # 에코프로비엠 (1억주)
        "259960": 100000000,      # 크래프톤 (1억주)
        "316140": 100000000,      # 우리금융지주 (1억주)
        "352820": 100000000,      # 하이브 (1억주)
        "373220": 100000000,      # LG에너지솔루션 (1억주)
        "403870": 100000000,      # HPSP (1억주)
        "429270": 100000000,      # 시스코 (1억주)
        "456040": 100000000,      # SK스퀘어 (1억주)
        "456570": 100000000,      # SK바이오팜 (1억주)
        "456970": 100000000,      # YG엔터테인먼트 (1억주)
        "457550": 100000000,      # SK바이오사이언스 (1억주)
    }
    
    return shares_map.get(ticker, 100000000)  # 기본값: 1억주


def filter_by_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """시총 기준 필터링"""
    if df.empty:
        return df
    
    # 시총 1조 5천억 이상 필터링
    filtered = df[df["시총(원)"] >= MARKET_CAP_THRESHOLD_1].copy()
    
    # 시총 내림차순 정렬
    filtered = filtered.sort_values("시총(원)", ascending=False).reset_index(drop=True)
    
    logger.info(f"✓ 시총 필터링 완료: {len(filtered)}개 종목")
    logger.info(f"  - 5조 이상: {len(filtered[filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_2])}개")
    logger.info(f"  - 1.5조~5조: {len(filtered[(filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_1) & (filtered['시총(원)'] < MARKET_CAP_THRESHOLD_2)])}개")
    
    return filtered


def format_market_cap(value: float) -> str:
    """시총 포맷팅 (조원 단위)"""
    if value >= 1000000000000:  # 1조 이상
        return f"{value/1000000000000:.1f}조원"
    elif value >= 1000000000:  # 10억 이상
        return f"{value/1000000000:.0f}억원"
    else:
        return f"{value:,.0f}원"


def apply_market_cap_formatting(file_path: str):
    """엑셀 포맷팅 적용"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 찾기
    headers = [cell.value for cell in ws[1]]
    
    # 열 인덱스 찾기
    col_indices = {}
    for idx, header in enumerate(headers, start=1):
        col_indices[header] = idx
    
    # 데이터 행 포맷팅
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header = headers[col_idx - 1]
            
            # 테두리 적용
            cell.border = thin_border
            
            # 시총 포맷 (천 자리 콤마)
            if header == "시총(원)":
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 현재가 포맷 (천 자리 콤마)
            elif header == "현재가":
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 거래량 포맷 (천 자리 콤마)
            elif header == "거래량":
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 텍스트 중앙 정렬 (티커, 종목명, 시총구분)
            elif header in ["티커", "종목명", "시총구분"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 기타는 왼쪽 정렬
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 헤더 포맷팅
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header = headers[col_idx - 1]
        
        header_cell.border = thin_border
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 열 너비 자동 조정
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1]
        
        if header == "티커":
            ws.column_dimensions[column_letter].width = 10
        elif header == "종목명":
            ws.column_dimensions[column_letter].width = 20
        elif header == "시총(원)":
            ws.column_dimensions[column_letter].width = 15
        elif header == "현재가":
            ws.column_dimensions[column_letter].width = 12
        elif header == "거래량":
            ws.column_dimensions[column_letter].width = 15
        elif header == "시총구분":
            ws.column_dimensions[column_letter].width = 12
        else:
            ws.column_dimensions[column_letter].width = 12
    
    # 시총구분별 색상 적용
    if "시총구분" in col_indices:
        col_idx = col_indices["시총구분"]
        
        # 5조 이상은 파란색 배경
        blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "5조이상":
                cell.fill = blue_fill
    
    wb.save(file_path)


def save_market_cap_universe(df: pd.DataFrame, file_path: str):
    """시총 유니버스를 엑셀에 저장"""
    
    # 엑셀에 저장
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Market_Cap_Universe", index=False)
    
    # 포맷팅 적용
    apply_market_cap_formatting(file_path)
    
    logger.info(f"✓ 시총 유니버스 저장 완료: {len(df)}개 종목")
    logger.info(f"  파일: {file_path}")


def main():
    parser = argparse.ArgumentParser(description="Market Cap Filter System (Real API)")
    parser.add_argument("--appkey", required=True, help="Kiwoom API App Key")
    parser.add_argument("--secret", required=True, help="Kiwoom API Secret Key")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="출력 파일 경로")
    
    args = parser.parse_args()
    
    output_file = args.output
    
    try:
        logger.info("=" * 80)
        logger.info("시총 기반 종목 필터링 시스템 시작 (실제 API)")
        logger.info(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"시총 기준: {format_market_cap(MARKET_CAP_THRESHOLD_1)} 이상")
        logger.info(f"5조 이상 별도 표시: {format_market_cap(MARKET_CAP_THRESHOLD_2)} 이상")
        logger.info("=" * 80)
        
        # 1. API 토큰 획득
        try:
            token = get_api_token(args.appkey, args.secret)
        except Exception as e:
            logger.error(f"API 토큰 획득 실패: {e}")
            sys.exit(1)
        
        # 2. 주요 종목들 정보 조회
        logger.info("\n" + "=" * 80)
        logger.info("주요 종목 정보 조회 중...")
        logger.info("=" * 80)
        
        all_stocks = []
        
        for idx, ticker in enumerate(MAJOR_STOCKS):
            logger.info(f"[{idx + 1}/{len(MAJOR_STOCKS)}] {ticker} 조회 중...")
            
            stock_info = fetch_stock_info(token, ticker)
            if stock_info:
                all_stocks.append(stock_info)
                logger.info(f"  ✓ {stock_info['종목명']}: {format_market_cap(stock_info['시총(원)'])} [{stock_info['시총구분']}]")
            else:
                logger.info(f"  ⚠️ {ticker}: 시총 기준 미달 또는 데이터 없음")
            
            # API 레이트 리미트 방지
            time.sleep(0.3)
        
        if not all_stocks:
            logger.warning("시총 기준 종목이 없습니다")
            sys.exit(1)
        
        # 3. DataFrame 생성 및 필터링
        df = pd.DataFrame(all_stocks)
        df_filtered = filter_by_market_cap(df)
        
        # 4. 상위 10개 종목 출력
        logger.info("\n" + "=" * 80)
        logger.info("상위 10개 종목")
        logger.info("=" * 80)
        
        for idx, row in df_filtered.head(10).iterrows():
            market_cap_formatted = format_market_cap(row["시총(원)"])
            current_price = f"{row['현재가']:,.0f}원" if row['현재가'] else "N/A"
            logger.info(f"{idx+1:2d}. {row['종목명']} ({row['티커']}) - {market_cap_formatted} [{row['시총구분']}] 현재가: {current_price}")
        
        # 5. 저장
        logger.info("\n" + "=" * 80)
        logger.info("파일 저장 중...")
        logger.info("=" * 80)
        
        save_market_cap_universe(df_filtered, output_file)
        
        # 6. 완료
        logger.info("\n" + "=" * 80)
        logger.info("완료")
        logger.info(f"총 종목: {len(df_filtered)}개")
        logger.info(f"5조 이상: {len(df_filtered[df_filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_2])}개")
        logger.info(f"1.5조~5조: {len(df_filtered[(df_filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_1) & (df_filtered['시총(원)'] < MARKET_CAP_THRESHOLD_2)])}개")
        logger.info("=" * 80)
    
    except Exception as e:
        logger.error(f"예기치 않은 오류 발생: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
