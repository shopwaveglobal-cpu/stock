"""
시총 기반 종목 필터링 시스템
- 매일 시총 1조 5천억 이상 종목들을 가져옴
- 5조 이상은 별도 표시
- market_cap_universe.xlsx 파일 생성
"""

import argparse
import logging
import time
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# API 설정
API_BASE_URL = "https://api.kiwoom.com"
API_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
API_STOCK_LIST_ENDPOINT = "/api/dostk/stock-list"
API_STOCK_LIST_ID = "ka10001"

# 기본 파일 경로
DEFAULT_OUTPUT_FILE = "output/market_cap_universe.xlsx"
MARKET_CAP_THRESHOLD_1 = 1500000000000  # 1조 5천억원
MARKET_CAP_THRESHOLD_2 = 5000000000000  # 5조원


def get_api_token(appkey: str, secret: str, max_retry: int = 3) -> str:
    """API 토큰 획득"""
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    body = {
        "grant_type": "client_credentials",
        "appkey": appkey,
        "secretkey": secret
    }
    
    for attempt in range(max_retry):
        try:
            response = requests.post(API_TOKEN_URL, headers=headers, json=body, timeout=20)
            response.raise_for_status()
            data = response.json()
            token = data.get("token") or data.get("access_token")
            
            if not token:
                raise ValueError("토큰을 찾을 수 없습니다")
            
            logger.info("✓ API 토큰 획득 완료")
            return token
            
        except Exception as e:
            if attempt == max_retry - 1:
                logger.error(f"토큰 획득 실패: {e}")
                raise
            logger.warning(f"토큰 획득 재시도 {attempt + 1}/{max_retry}")
            time.sleep(1)
    
    raise RuntimeError("토큰 획득 실패")


def fetch_stock_list(token: str, max_retry: int = 5) -> pd.DataFrame:
    """전체 종목 리스트 조회 (KRX+NXT 통합 기준)"""
    headers = {
        "authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "api-id": API_STOCK_LIST_ID,
        "cont-yn": "N",
        "next-key": ""
    }
    
    body = {
        "stex_tp": "3",  # 통합 (KRX+NXT)
        "stk_cd": "",    # 전체 종목
        "stk_nm": "",    # 전체 종목명
        "page_no": "1",
        "page_cnt": "1000"  # 한 번에 1000개씩
    }
    
    url = API_BASE_URL + API_STOCK_LIST_ENDPOINT
    all_stocks = []
    page_no = 1
    
    while True:
        body["page_no"] = str(page_no)
        
        for attempt in range(max_retry):
            try:
                response = requests.post(url, headers=headers, json=body, timeout=20)
                response.raise_for_status()
                result = response.json()
                
                # 데이터 추출
                records = None
                for value in result.values():
                    if isinstance(value, list) and len(value) > 0:
                        records = value
                        break
                
                if not records:
                    logger.info(f"  📄 페이지 {page_no}: 데이터 없음 - 종료")
                    return pd.DataFrame(all_stocks)
                
                logger.info(f"  📄 페이지 {page_no}: {len(records)}개 종목 조회")
                
                # 데이터 파싱 (KRX+NXT 통합 기준)
                for rec in records:
                    try:
                        # 종목코드 (통합 기준에서는 _AL 접미사가 붙을 수 있음)
                        ticker = rec.get("stk_cd") or rec.get("ticker")
                        if not ticker:
                            continue
                        
                        # _AL 접미사 제거 (원본 티커만 사용)
                        if ticker.endswith("_AL"):
                            ticker = ticker[:-3]
                        
                        # 종목명
                        name = rec.get("stk_nm") or rec.get("name")
                        if not name:
                            continue
                        
                        # 시총 (단위: 원) - 통합 기준에서는 다른 필드명 사용 가능
                        market_cap_str = rec.get("mkt_cap") or rec.get("market_cap") or rec.get("mkt_cap_val") or rec.get("tot_mkt_cap")
                        if not market_cap_str:
                            continue
                        
                        try:
                            # 콤마 제거 후 숫자 변환
                            market_cap = float(str(market_cap_str).replace(",", ""))
                        except (ValueError, TypeError):
                            continue
                        
                        # 현재가 (통합 기준)
                        current_price_str = rec.get("cur_prc") or rec.get("current_price") or rec.get("stck_prpr") or rec.get("prc")
                        current_price = None
                        if current_price_str:
                            try:
                                current_price = float(str(current_price_str).replace(",", ""))
                            except (ValueError, TypeError):
                                pass
                        
                        # 거래량 (통합 기준)
                        volume_str = rec.get("acml_vol") or rec.get("volume") or rec.get("trd_vol") or rec.get("vol")
                        volume = None
                        if volume_str:
                            try:
                                volume = int(str(volume_str).replace(",", ""))
                            except (ValueError, TypeError):
                                pass
                        
                        all_stocks.append({
                            "티커": ticker,
                            "종목명": name,
                            "시총(원)": market_cap,
                            "현재가": current_price,
                            "거래량": volume,
                            "시총구분": "5조이상" if market_cap >= MARKET_CAP_THRESHOLD_2 else "1.5조이상"
                        })
                        
                    except Exception as e:
                        logger.warning(f"  ⚠️ 종목 데이터 파싱 실패: {e}")
                        continue
                
                # 다음 페이지로
                page_no += 1
                break
                
            except Exception as e:
                if attempt == max_retry - 1:
                    logger.error(f"  ❌ 페이지 {page_no} 조회 실패: {e}")
                    return pd.DataFrame(all_stocks)
                time.sleep(0.5 * (2 ** attempt))
        
        # API 레이트 리미트 방지
        time.sleep(0.2)
    
    return pd.DataFrame(all_stocks)


def filter_by_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """시총 기준 필터링"""
    if df.empty:
        return df
    
    # 시총 1조 5천억 이상 필터링
    filtered = df[df["시총(원)"] >= MARKET_CAP_THRESHOLD_1].copy()
    
    # 시총 내림차순 정렬
    filtered = filtered.sort_values("시총(원)", ascending=False).reset_index(drop=True)
    
    logger.info(f"✓ 시총 필터링 완료: {len(filtered)}개 종목")
    logger.info(f"  - 5조 이상: {len(filtered[filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_2])}개")
    logger.info(f"  - 1.5조~5조: {len(filtered[(filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_1) & (filtered['시총(원)'] < MARKET_CAP_THRESHOLD_2)])}개")
    
    return filtered


def format_market_cap(value: float) -> str:
    """시총 포맷팅 (조원 단위)"""
    if value >= 1000000000000:  # 1조 이상
        return f"{value/1000000000000:.1f}조원"
    elif value >= 1000000000:  # 10억 이상
        return f"{value/1000000000:.0f}억원"
    else:
        return f"{value:,.0f}원"


def apply_market_cap_formatting(file_path: str):
    """엑셀 포맷팅 적용"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 찾기
    headers = [cell.value for cell in ws[1]]
    
    # 열 인덱스 찾기
    col_indices = {}
    for idx, header in enumerate(headers, start=1):
        col_indices[header] = idx
    
    # 데이터 행 포맷팅
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header = headers[col_idx - 1]
            
            # 테두리 적용
            cell.border = thin_border
            
            # 시총 포맷 (천 자리 콤마)
            if header == "시총(원)":
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 현재가 포맷 (천 자리 콤마)
            elif header == "현재가":
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 거래량 포맷 (천 자리 콤마)
            elif header == "거래량":
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 텍스트 중앙 정렬 (티커, 종목명, 시총구분)
            elif header in ["티커", "종목명", "시총구분"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 기타는 왼쪽 정렬
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 헤더 포맷팅
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header = headers[col_idx - 1]
        
        header_cell.border = thin_border
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 열 너비 자동 조정
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1]
        
        if header == "티커":
            ws.column_dimensions[column_letter].width = 10
        elif header == "종목명":
            ws.column_dimensions[column_letter].width = 20
        elif header == "시총(원)":
            ws.column_dimensions[column_letter].width = 15
        elif header == "현재가":
            ws.column_dimensions[column_letter].width = 12
        elif header == "거래량":
            ws.column_dimensions[column_letter].width = 15
        elif header == "시총구분":
            ws.column_dimensions[column_letter].width = 12
        else:
            ws.column_dimensions[column_letter].width = 12
    
    # 시총구분별 색상 적용
    if "시총구분" in col_indices:
        col_idx = col_indices["시총구분"]
        
        # 5조 이상은 파란색 배경
        blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "5조이상":
                cell.fill = blue_fill
    
    wb.save(file_path)


def save_market_cap_universe(df: pd.DataFrame, file_path: str):
    """시총 유니버스를 엑셀에 저장"""
    
    # 엑셀에 저장
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Market_Cap_Universe", index=False)
    
    # 포맷팅 적용
    apply_market_cap_formatting(file_path)
    
    logger.info(f"✓ 시총 유니버스 저장 완료: {len(df)}개 종목")
    logger.info(f"  파일: {file_path}")


def main():
    parser = argparse.ArgumentParser(description="Market Cap Filter System")
    parser.add_argument("--appkey", required=True, help="Kiwoom API App Key")
    parser.add_argument("--secret", required=True, help="Kiwoom API Secret Key")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="출력 파일 경로")
    
    args = parser.parse_args()
    
    output_file = args.output
    
    try:
        logger.info("=" * 80)
        logger.info("시총 기반 종목 필터링 시스템 시작")
        logger.info(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"시총 기준: {format_market_cap(MARKET_CAP_THRESHOLD_1)} 이상")
        logger.info(f"5조 이상 별도 표시: {format_market_cap(MARKET_CAP_THRESHOLD_2)} 이상")
        logger.info("=" * 80)
        
        # 1. API 토큰 획득
        try:
            token = get_api_token(args.appkey, args.secret)
        except Exception as e:
            logger.error(f"API 토큰 획득 실패: {e}")
            sys.exit(1)
        
        # 2. 전체 종목 리스트 조회
        logger.info("\n" + "=" * 80)
        logger.info("전체 종목 리스트 조회 중...")
        logger.info("=" * 80)
        
        df_all_stocks = fetch_stock_list(token)
        
        if df_all_stocks.empty:
            logger.error("종목 리스트 조회 실패")
            sys.exit(1)
        
        logger.info(f"✓ 전체 종목 조회 완료: {len(df_all_stocks)}개")
        
        # 3. 시총 기준 필터링
        logger.info("\n" + "=" * 80)
        logger.info("시총 기준 필터링 중...")
        logger.info("=" * 80)
        
        df_filtered = filter_by_market_cap(df_all_stocks)
        
        if df_filtered.empty:
            logger.warning("시총 기준 종목이 없습니다")
            sys.exit(1)
        
        # 4. 상위 10개 종목 출력
        logger.info("\n" + "=" * 80)
        logger.info("상위 10개 종목")
        logger.info("=" * 80)
        
        for idx, row in df_filtered.head(10).iterrows():
            market_cap_formatted = format_market_cap(row["시총(원)"])
            current_price = f"{row['현재가']:,.0f}원" if row['현재가'] else "N/A"
            logger.info(f"{idx+1:2d}. {row['종목명']} ({row['티커']}) - {market_cap_formatted} [{row['시총구분']}] 현재가: {current_price}")
        
        # 5. 저장
        logger.info("\n" + "=" * 80)
        logger.info("파일 저장 중...")
        logger.info("=" * 80)
        
        save_market_cap_universe(df_filtered, output_file)
        
        # 6. 완료
        logger.info("\n" + "=" * 80)
        logger.info("완료")
        logger.info(f"총 종목: {len(df_filtered)}개")
        logger.info(f"5조 이상: {len(df_filtered[df_filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_2])}개")
        logger.info(f"1.5조~5조: {len(df_filtered[(df_filtered['시총(원)'] >= MARKET_CAP_THRESHOLD_1) & (df_filtered['시총(원)'] < MARKET_CAP_THRESHOLD_2)])}개")
        logger.info("=" * 80)
    
    except Exception as e:
        logger.error(f"예기치 않은 오류 발생: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
