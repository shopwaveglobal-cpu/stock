"""
KRX 데이터를 활용한 시총 기반 종목 필터링 시스템 (최종 버전)
pykrx를 사용해서 전체 종목의 시가총액을 한 번에 가져와서 필터링
"""

import argparse
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd

# pykrx 라이브러리 임포트
try:
    from pykrx import stock
    HAS_PYKRX = True
except ImportError:
    HAS_PYKRX = False
    print("pykrx 라이브러리가 설치되지 않았습니다.")
    print("설치 방법: pip install pykrx")
    sys.exit(1)

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# 시총 기준 (원 단위)
MARKET_CAP_THRESHOLD_1 = 1_500_000_000_000  # 1.5조원
MARKET_CAP_THRESHOLD_2 = 5_000_000_000_000  # 5조원

# 출력 파일
DEFAULT_OUTPUT_FILE = "output/market_cap_universe.xlsx"


def get_krx_market_cap_data(date: str = None) -> pd.DataFrame:
    """
    KRX에서 전체 종목의 시가총액 데이터 가져오기
    
    Args:
        date: 조회 날짜 (YYYYMMDD), None이면 최근 영업일
    
    Returns:
        DataFrame with columns: 티커, 종목명, 시가총액, 현재가, 거래량 등
    """
    logger.info("KRX에서 시가총액 데이터 가져오는 중...")
    
    try:
        # 날짜 설정 (어제 날짜 사용 - 시가총액 데이터 안정성)
        if date is None:
            # 어제 날짜 사용 (시가총액 데이터가 안정적으로 업데이트됨)
            yesterday = datetime.now() - timedelta(days=1)
            date = yesterday.strftime("%Y%m%d")
            logger.info(f"  어제 날짜 사용: {date}")
        
        # KOSPI 종목 가져오기
        logger.info("  KOSPI 종목 조회 중...")
        try:
            kospi_df = stock.get_market_cap_by_ticker(date, market="KOSPI")
            kospi_df['시장'] = 'KOSPI'
            logger.info(f"    KOSPI: {len(kospi_df)}개 종목")
        except Exception as e:
            logger.warning(f"    KOSPI 조회 실패: {e}")
            kospi_df = pd.DataFrame()
        
        # KOSDAQ 종목 가져오기
        logger.info("  KOSDAQ 종목 조회 중...")
        try:
            kosdaq_df = stock.get_market_cap_by_ticker(date, market="KOSDAQ")
            kosdaq_df['시장'] = 'KOSDAQ'
            logger.info(f"    KOSDAQ: {len(kosdaq_df)}개 종목")
        except Exception as e:
            logger.warning(f"    KOSDAQ 조회 실패: {e}")
            kosdaq_df = pd.DataFrame()
        
        # 종목명 추가
        logger.info("  종목명 조회 중...")
        try:
            # KOSPI 종목명
            if not kospi_df.empty:
                kospi_tickers = kospi_df.index.tolist()
                kospi_names = []
                for ticker in kospi_tickers:
                    try:
                        name = stock.get_market_ticker_name(ticker)
                        kospi_names.append(name if name else f"종목_{ticker}")
                    except:
                        kospi_names.append(f"종목_{ticker}")
                kospi_df['종목명'] = kospi_names
                logger.info(f"    KOSPI 종목명 {len(kospi_names)}개 조회 완료")
            
            # KOSDAQ 종목명
            if not kosdaq_df.empty:
                kosdaq_tickers = kosdaq_df.index.tolist()
                kosdaq_names = []
                for ticker in kosdaq_tickers:
                    try:
                        name = stock.get_market_ticker_name(ticker)
                        kosdaq_names.append(name if name else f"종목_{ticker}")
                    except:
                        kosdaq_names.append(f"종목_{ticker}")
                kosdaq_df['종목명'] = kosdaq_names
                logger.info(f"    KOSDAQ 종목명 {len(kosdaq_names)}개 조회 완료")
        except Exception as e:
            logger.warning(f"    종목명 조회 실패: {e}")
            # 종목명이 없으면 티커로 대체
            if not kospi_df.empty:
                kospi_df['종목명'] = kospi_df.index
            if not kosdaq_df.empty:
                kosdaq_df['종목명'] = kosdaq_df.index
        
        # 합치기
        if not kospi_df.empty and not kosdaq_df.empty:
            all_df = pd.concat([kospi_df, kosdaq_df], axis=0)
        elif not kospi_df.empty:
            all_df = kospi_df
        elif not kosdaq_df.empty:
            all_df = kosdaq_df
        else:
            logger.error("모든 시장에서 데이터를 가져올 수 없습니다.")
            return pd.DataFrame()
        
        # 인덱스를 티커로 변환
        all_df = all_df.reset_index()
        all_df = all_df.rename(columns={
            '티커': '티커',
            '종목명': '종목명',
            '시가총액': '시총(원)',
            '종가': '현재가',
            '거래량': '거래량',
            '거래대금': '거래대금',
            '상장주식수': '상장주식수'
        })
        
        # 컬럼명 확인 및 조정
        if '티커' not in all_df.columns and all_df.index.name == '티커':
            all_df = all_df.reset_index()
        
        logger.info(f"  총 {len(all_df)}개 종목 데이터 수집 완료")
        
        return all_df
        
    except Exception as e:
        logger.error(f"KRX 데이터 가져오기 실패: {e}")
        logger.error("최근 영업일 데이터를 시도합니다...")
        
        # 최근 영업일 시도 (어제)
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
        
        try:
            logger.info(f"  어제 날짜({yesterday})로 재시도...")
            kospi_df = stock.get_market_cap_by_ticker(yesterday, market="KOSPI")
            kospi_df['시장'] = 'KOSPI'
            kosdaq_df = stock.get_market_cap_by_ticker(yesterday, market="KOSDAQ")
            kosdaq_df['시장'] = 'KOSDAQ'
            all_df = pd.concat([kospi_df, kosdaq_df], axis=0)
            all_df = all_df.reset_index()
            
            logger.info(f"  어제 날짜로 {len(all_df)}개 종목 데이터 수집 완료")
            return all_df
            
        except Exception as e2:
            logger.error(f"어제 데이터도 실패: {e2}")
            raise


def filter_by_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """시총 1.5조 이상 종목 필터링"""
    logger.info(f"\n시총 필터링 중... (기준: 1.5조원 이상)")
    
    # 시총 컬럼 확인
    market_cap_col = None
    for col in ['시총(원)', '시가총액', 'market_cap']:
        if col in df.columns:
            market_cap_col = col
            break
    
    if market_cap_col is None:
        logger.error(f"시가총액 컬럼을 찾을 수 없습니다. 현재 컬럼: {df.columns.tolist()}")
        return pd.DataFrame()
    
    # 시총 1.5조 이상 필터링
    filtered = df[df[market_cap_col] >= MARKET_CAP_THRESHOLD_1].copy()
    
    # 시총구분 추가
    filtered['시총구분'] = filtered[market_cap_col].apply(
        lambda x: "5조이상" if x >= MARKET_CAP_THRESHOLD_2 else "1.5조이상"
    )
    
    # 시총 내림차순 정렬
    filtered = filtered.sort_values(market_cap_col, ascending=False).reset_index(drop=True)
    
    logger.info(f"  필터링 결과: {len(filtered)}개 종목")
    logger.info(f"  - 5조 이상: {len(filtered[filtered['시총구분'] == '5조이상'])}개")
    logger.info(f"  - 1.5조~5조: {len(filtered[filtered['시총구분'] == '1.5조이상'])}개")
    
    return filtered


def format_market_cap(value: float) -> str:
    """시총 포맷팅 (조원 단위)"""
    if value >= 1_000_000_000_000:  # 1조 이상
        return f"{value/1_000_000_000_000:.1f}조원"
    elif value >= 100_000_000:  # 1억 이상
        return f"{value/100_000_000:.0f}억원"
    else:
        return f"{value:,.0f}원"


def save_to_excel(df: pd.DataFrame, file_path: str):
    """엑셀 파일로 저장"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    
    # 엑셀 저장
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Market_Cap_Universe", index=False)
    
    # 포맷팅
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 5조 이상 종목 색상
    blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    headers = [cell.value for cell in ws[1]]
    if '시총구분' in headers:
        col_idx = headers.index('시총구분') + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "5조이상":
                cell.fill = blue_fill
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 10  # 티커
    ws.column_dimensions['B'].width = 20  # 종목명
    ws.column_dimensions['C'].width = 18  # 시총
    ws.column_dimensions['D'].width = 12  # 현재가
    ws.column_dimensions['E'].width = 15  # 거래량
    ws.column_dimensions['F'].width = 10  # 시장
    ws.column_dimensions['G'].width = 12  # 시총구분
    
    wb.save(file_path)
    logger.info(f"  파일 저장 완료: {file_path}")


def main():
    parser = argparse.ArgumentParser(description="KRX 데이터 기반 시총 필터링 시스템 (pykrx)")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="출력 파일 경로")
    parser.add_argument("--date", default=None, help="조회 날짜 (YYYYMMDD)")
    
    args = parser.parse_args()
    
    try:
        logger.info("=" * 80)
        logger.info("KRX 데이터 기반 시총 필터링 시스템 시작 (pykrx)")
        logger.info(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"시총 기준: {format_market_cap(MARKET_CAP_THRESHOLD_1)} 이상")
        logger.info("=" * 80)
        
        # 1. KRX에서 시가총액 데이터 가져오기
        df = get_krx_market_cap_data(args.date)
        
        if df.empty:
            logger.warning("데이터를 가져오지 못했습니다")
            sys.exit(1)
        
        # 2. 시총 필터링
        df_filtered = filter_by_market_cap(df)
        
        if df_filtered.empty:
            logger.warning("필터링된 종목이 없습니다")
            sys.exit(1)
        
        # 3. 상위 20개 종목 출력
        logger.info("\n" + "=" * 80)
        logger.info("상위 20개 종목")
        logger.info("=" * 80)
        
        for idx, row in df_filtered.head(20).iterrows():
            ticker = row.get('티커', row.get('종목코드', 'N/A'))
            name = row.get('종목명', 'N/A')
            market_cap_col = '시총(원)' if '시총(원)' in row else '시가총액'
            market_cap = row.get(market_cap_col, 0)
            market_cap_formatted = format_market_cap(market_cap)
            category = row.get('시총구분', 'N/A')
            market = row.get('시장', 'N/A')
            
            logger.info(f"{idx+1:2d}. {name} ({ticker}) - {market_cap_formatted} [{category}] {market}")
        
        # 4. 엑셀 저장
        logger.info("\n" + "=" * 80)
        logger.info("엑셀 파일 저장 중...")
        logger.info("=" * 80)
        
        # 출력 디렉토리 생성
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        
        # 필요한 컬럼만 선택
        output_columns = []
        if '티커' in df_filtered.columns:
            output_columns.append('티커')
        elif '종목코드' in df_filtered.columns:
            df_filtered['티커'] = df_filtered['종목코드']
            output_columns.append('티커')
        
        if '종목명' in df_filtered.columns:
            output_columns.append('종목명')
        
        market_cap_col = '시총(원)' if '시총(원)' in df_filtered.columns else '시가총액'
        if market_cap_col in df_filtered.columns:
            if market_cap_col != '시총(원)':
                df_filtered['시총(원)'] = df_filtered[market_cap_col]
            output_columns.append('시총(원)')
        
        if '종가' in df_filtered.columns:
            df_filtered['현재가'] = df_filtered['종가']
            output_columns.append('현재가')
        elif '현재가' in df_filtered.columns:
            output_columns.append('현재가')
        
        if '거래량' in df_filtered.columns:
            output_columns.append('거래량')
        
        if '시장' in df_filtered.columns:
            output_columns.append('시장')
        
        if '시총구분' in df_filtered.columns:
            output_columns.append('시총구분')
        
        df_output = df_filtered[output_columns].copy()
        
        save_to_excel(df_output, args.output)
        
        logger.info("\n" + "=" * 80)
        logger.info("시총 필터링 완료!")
        logger.info(f"  총 {len(df_filtered)}개 종목")
        logger.info(f"  파일: {args.output}")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"오류 발생: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
