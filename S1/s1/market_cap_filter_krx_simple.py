"""
KRX 웹사이트에서 직접 시가총액 데이터 가져오기
pykrx 없이 requests만 사용
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
import requests
import json

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# 시총 기준 (원 단위)
MARKET_CAP_THRESHOLD_1 = 1_500_000_000_000  # 1.5조원
MARKET_CAP_THRESHOLD_2 = 5_000_000_000_000  # 5조원

# 출력 파일
DEFAULT_OUTPUT_FILE = "output/market_cap_universe.xlsx"


def get_krx_stock_data():
    """KRX 웹사이트에서 전체 종목 시가총액 데이터 가져오기"""
    logger.info("KRX에서 시가총액 데이터 가져오는 중...")
    
    # KRX 정보데이터시스템 API
    url = "http://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd"
    
    # 오늘 날짜
    today = datetime.now().strftime("%Y%m%d")
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "http://data.krx.co.kr/contents/MDC/MDI/mdiLoader",
    }
    
    all_stocks = []
    
    # KOSPI 데이터 가져오기
    logger.info("  KOSPI 종목 조회 중...")
    try:
        kospi_data = {
            "bld": "dbms/MDC/STAT/standard/MDCSTAT01501",
            "locale": "ko_KR",
            "mktId": "STK",  # KOSPI
            "trdDd": today,
            "share": "1",
            "money": "1",
            "csvxls_isNo": "false",
        }
        
        response = requests.post(url, data=kospi_data, headers=headers, timeout=30)
        response.raise_for_status()
        result = response.json()
        
        if 'OutBlock_1' in result:
            kospi_stocks = result['OutBlock_1']
            logger.info(f"    KOSPI: {len(kospi_stocks)}개 종목")
            
            for stock in kospi_stocks:
                try:
                    ticker = stock.get('ISU_SRT_CD', '').strip()
                    name = stock.get('ISU_ABBRV', '').strip()
                    market_cap = int(stock.get('MKTCAP', '0').replace(',', ''))  # 백만원 단위
                    market_cap_won = market_cap * 1_000_000  # 원 단위로 변환
                    close_price = int(stock.get('TDD_CLSPRC', '0').replace(',', ''))
                    volume = int(stock.get('ACC_TRDVOL', '0').replace(',', ''))
                    
                    if ticker and market_cap_won >= MARKET_CAP_THRESHOLD_1:
                        all_stocks.append({
                            '티커': ticker,
                            '종목명': name,
                            '시총(원)': market_cap_won,
                            '현재가': close_price,
                            '거래량': volume,
                            '시장': 'KOSPI',
                            '시총구분': "5조이상" if market_cap_won >= MARKET_CAP_THRESHOLD_2 else "1.5조이상"
                        })
                except Exception as e:
                    logger.warning(f"    KOSPI 종목 파싱 실패: {e}")
                    continue
        
    except Exception as e:
        logger.error(f"  KOSPI 데이터 조회 실패: {e}")
    
    # KOSDAQ 데이터 가져오기
    logger.info("  KOSDAQ 종목 조회 중...")
    try:
        kosdaq_data = {
            "bld": "dbms/MDC/STAT/standard/MDCSTAT01501",
            "locale": "ko_KR",
            "mktId": "KSQ",  # KOSDAQ
            "trdDd": today,
            "share": "1",
            "money": "1",
            "csvxls_isNo": "false",
        }
        
        response = requests.post(url, data=kosdaq_data, headers=headers, timeout=30)
        response.raise_for_status()
        result = response.json()
        
        if 'OutBlock_1' in result:
            kosdaq_stocks = result['OutBlock_1']
            logger.info(f"    KOSDAQ: {len(kosdaq_stocks)}개 종목")
            
            for stock in kosdaq_stocks:
                try:
                    ticker = stock.get('ISU_SRT_CD', '').strip()
                    name = stock.get('ISU_ABBRV', '').strip()
                    market_cap = int(stock.get('MKTCAP', '0').replace(',', ''))  # 백만원 단위
                    market_cap_won = market_cap * 1_000_000  # 원 단위로 변환
                    close_price = int(stock.get('TDD_CLSPRC', '0').replace(',', ''))
                    volume = int(stock.get('ACC_TRDVOL', '0').replace(',', ''))
                    
                    if ticker and market_cap_won >= MARKET_CAP_THRESHOLD_1:
                        all_stocks.append({
                            '티커': ticker,
                            '종목명': name,
                            '시총(원)': market_cap_won,
                            '현재가': close_price,
                            '거래량': volume,
                            '시장': 'KOSDAQ',
                            '시총구분': "5조이상" if market_cap_won >= MARKET_CAP_THRESHOLD_2 else "1.5조이상"
                        })
                except Exception as e:
                    logger.warning(f"    KOSDAQ 종목 파싱 실패: {e}")
                    continue
        
    except Exception as e:
        logger.error(f"  KOSDAQ 데이터 조회 실패: {e}")
    
    logger.info(f"  총 {len(all_stocks)}개 종목 (시총 1.5조 이상) 수집 완료")
    
    return pd.DataFrame(all_stocks)


def format_market_cap(value: float) -> str:
    """시총 포맷팅 (조원 단위)"""
    if value >= 1_000_000_000_000:  # 1조 이상
        return f"{value/1_000_000_000_000:.1f}조원"
    elif value >= 100_000_000:  # 1억 이상
        return f"{value/100_000_000:.0f}억원"
    else:
        return f"{value:,.0f}원"


def save_to_excel(df: pd.DataFrame, file_path: str):
    """엑셀 파일로 저장"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    
    # 엑셀 저장
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Market_Cap_Universe", index=False)
    
    # 포맷팅
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 5조 이상 종목 색상
    blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    headers = [cell.value for cell in ws[1]]
    if '시총구분' in headers:
        col_idx = headers.index('시총구분') + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == "5조이상":
                cell.fill = blue_fill
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 10  # 티커
    ws.column_dimensions['B'].width = 20  # 종목명
    ws.column_dimensions['C'].width = 18  # 시총
    ws.column_dimensions['D'].width = 12  # 현재가
    ws.column_dimensions['E'].width = 15  # 거래량
    ws.column_dimensions['F'].width = 10  # 시장
    ws.column_dimensions['G'].width = 12  # 시총구분
    
    wb.save(file_path)
    logger.info(f"  파일 저장 완료: {file_path}")


def main():
    parser = argparse.ArgumentParser(description="KRX 데이터 기반 시총 필터링 시스템 (Simple)")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="출력 파일 경로")
    
    args = parser.parse_args()
    
    try:
        logger.info("=" * 80)
        logger.info("KRX 데이터 기반 시총 필터링 시스템 시작")
        logger.info(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"시총 기준: {format_market_cap(MARKET_CAP_THRESHOLD_1)} 이상")
        logger.info("=" * 80)
        
        # 1. KRX에서 시가총액 데이터 가져오기
        df = get_krx_stock_data()
        
        if df.empty:
            logger.warning("데이터를 가져오지 못했습니다")
            sys.exit(1)
        
        # 2. 시총 내림차순 정렬
        df = df.sort_values('시총(원)', ascending=False).reset_index(drop=True)
        
        logger.info(f"\n필터링 결과: {len(df)}개 종목")
        logger.info(f"  - 5조 이상: {len(df[df['시총구분'] == '5조이상'])}개")
        logger.info(f"  - 1.5조~5조: {len(df[df['시총구분'] == '1.5조이상'])}개")
        
        # 3. 상위 20개 종목 출력
        logger.info("\n" + "=" * 80)
        logger.info("상위 20개 종목")
        logger.info("=" * 80)
        
        for idx, row in df.head(20).iterrows():
            market_cap_formatted = format_market_cap(row['시총(원)'])
            logger.info(f"{idx+1:2d}. {row['종목명']} ({row['티커']}) - {market_cap_formatted} [{row['시총구분']}] {row['시장']}")
        
        # 4. 엑셀 저장
        logger.info("\n" + "=" * 80)
        logger.info("엑셀 파일 저장 중...")
        logger.info("=" * 80)
        
        # 출력 디렉토리 생성
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        
        save_to_excel(df, args.output)
        
        logger.info("\n" + "=" * 80)
        logger.info("시총 필터링 완료!")
        logger.info(f"  총 {len(df)}개 종목")
        logger.info(f"  파일: {args.output}")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"오류 발생: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()


