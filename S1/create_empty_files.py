#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""빈 Excel 파일 생성 (테스트용)"""

from openpyxl import Workbook
import os

# output 디렉토리 확인
os.makedirs("output", exist_ok=True)

# 1. marketcap_universe.xlsx 생성
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "universe"
ws1.append(["첫주도주", "최근주도주", "티커", "종목명", "시가총액(억)", "누적횟수"])
ws1["H1"] = "최종 업데이트: 2025-11-02"
wb1.save("output/marketcap_universe.xlsx")
print("marketcap_universe.xlsx created")

# 2. trading_signals_s1.xlsx 생성
wb2 = Workbook()
ws2_summary = wb2.active
ws2_summary.title = "Summary"
ws2_summary.append(["티커", "종목명", "매수상태", "알람상태"])

ws2_history = wb2.create_sheet("History")
ws2_history.append(["티커", "종목명", "매수상태"])

wb2.save("output/trading_signals_s1.xlsx")
print("trading_signals_s1.xlsx created")

